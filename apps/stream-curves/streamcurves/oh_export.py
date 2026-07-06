"""Port of R/14_oh_list_of_metrics.R and R/15_oh_sqt_workbook.R.

Output Handbook (OH/SQT) xlsx builders:

* :func:`build_oh_list_of_metrics` — loads the MN "List of Metrics" template
  (``data/templates/MN-List-of-Metricsv2.0.xlsx``), overwrites the data cells
  with project threshold values and preserves the "Notes" footer (row 55+).
* :func:`build_oh_reference_curves_workbook` — builds the "Reference
  Standards" workbook from scratch: five horizontal discipline bands, one
  stacked block per metric × stratum, plus a "Pull Down Notes" sheet.

openxlsx → openpyxl translation notes
-------------------------------------
* R passed ``(wb, sheet)`` pairs to the sheet writers; here they take an
  openpyxl ``Worksheet`` directly (``oh_clear_performance_standards_data``,
  ``oh_write_performance_standards``, ``oh_write_references_sheet``,
  ``oh_sqt_write_discipline_header_row``, ``oh_sqt_write_metric_block``).
* The R argument ``discipline_function_mapping`` is named ``mapping`` here
  (per the port spec). It is the same pandas DataFrame contract: columns
  ``metric_key, discipline, function_label, sort_order``.
* ``build_oh_reference_curves_workbook`` drops R's ``template_path`` argument
  — the R function accepted it but never used it (the workbook is built from
  scratch).
* ``oh_lom_strip_phantom_drawing_refs`` is NOT ported: it worked around
  openxlsx emitting dangling drawing relationship entries; openpyxl does not
  produce them.
* openxlsx silently writes into merged regions; openpyxl raises when writing
  a non-anchor ``MergedCell``. Template merged ranges intersecting the data
  regions are therefore unmerged before clearing (the R code populated every
  underlying cell precisely because the template merges don't line up with
  Ohio's rows, so unmerging matches the intent: every row shows its values).
* Plot images: R embedded ``context$metrics[[metric_key]]$plot_file`` (a
  ggplot PNG saved under ``bundle_dir``). The port supports that same
  ``plot_file`` + ``bundle_dir`` contract and, additionally, a
  ``plot_png`` entry per metric in ``context["metrics"]`` holding either raw
  PNG bytes or a zero-argument callable returning PNG bytes. When neither is
  available the image is skipped (everything else in the block is written).
  Images are sized 346x221 px (R: 3.6in x 2.3in at 96 dpi).

Internal mirrors
----------------
The parameter-map accessors and mapping resolvers come from
:mod:`streamcurves.mapping` (the R/13_oh_parameter_map.R port). Only the
curve-point readers from R/10_reference_curves.R are mirrored privately here
(prefixed ``_``) so this module does not depend on the curves module — the
``threshold_rows`` DataFrame is consumed purely through its columns.
"""

from __future__ import annotations

import logging
import math
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .mapping import (
    fixed_discipline_order,
    metrics_for_resolved_discipline,
    oh_covered_metrics,
    oh_data_sources,
    oh_metric_entry,
    oh_reference_notes,
    oh_units_display,
    resolve_metric_function,
    resolved_category_order,
)

logger = logging.getLogger("streamcurves")


# --------------------------------------------------------------------------- #
# Small shared helpers
# --------------------------------------------------------------------------- #


def _or(x, default):
    """R ``%||%``: default only when x is None."""
    return x if x is not None else default


def _is_na_scalar(x) -> bool:
    if x is None:
        return True
    try:
        return bool(pd.isna(x))
    except (TypeError, ValueError):
        return False


def _is_true(x) -> bool:
    """R ``isTRUE``: True only for a logical TRUE scalar (not 1, not NA)."""
    return isinstance(x, (bool, np.bool_)) and bool(x)


def _chr(v) -> str:
    """R/14's row_vec coercion: NULL/NA/"" -> "", else as.character."""
    if v is None or _is_na_scalar(v):
        return ""
    return str(v)


def _row_scalar(row: pd.DataFrame, col: str):
    """``row[[col]][1]`` with R semantics: missing column -> None (R NULL)."""
    if row is None or col not in getattr(row, "columns", ()) or len(row) == 0:
        return None
    return row[col].iloc[0]


def _signif(x: float, digits: int = 3) -> float:
    """R ``signif``: round to significant digits (half-to-even, like R)."""
    x = float(x)
    if x == 0 or not math.isfinite(x):
        return x
    return round(x, digits - 1 - math.floor(math.log10(abs(x))))


def _format_number(x: float) -> str:
    """R ``format(x, trim = TRUE, scientific = FALSE)`` for signif-rounded
    doubles: shortest positional decimal representation ("41", "0.00123",
    "1230000")."""
    return np.format_float_positional(float(x), trim="-")


def _has_stratum_label(stratum_label) -> bool:
    """R: !is.na(x) && nzchar(x) && !identical(x, "none")."""
    if _is_na_scalar(stratum_label):
        return False
    s = str(stratum_label)
    return bool(s) and s != "none"


# --------------------------------------------------------------------------- #
# Mapping plumbing (accessors/resolvers themselves live in streamcurves.mapping)
# --------------------------------------------------------------------------- #


def _mapping_is_usable(mapping) -> bool:
    """R's recurring guard: mapping is a non-empty data.frame."""
    return isinstance(mapping, pd.DataFrame) and len(mapping) > 0


# --------------------------------------------------------------------------- #
# Private mirrors of the curve-point readers from R/10_reference_curves.R
# --------------------------------------------------------------------------- #


def _empty_reference_curve_points() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "point_order": pd.Series([], dtype="int64"),
            "metric_value": pd.Series([], dtype="float64"),
            "index_score": pd.Series([], dtype="float64"),
        }
    )


def _normalize_reference_curve_points(curve_points) -> pd.DataFrame:
    if curve_points is None:
        return _empty_reference_curve_points()
    if isinstance(curve_points, pd.DataFrame):
        points = curve_points.copy()
    elif isinstance(curve_points, (list, dict)):
        points = pd.DataFrame(curve_points)
    else:
        raise ValueError("Curve points must be NULL, a data frame, or a list.")

    if "metric_value" not in points.columns and "x" in points.columns:
        points["metric_value"] = points["x"]
    if "index_score" not in points.columns and "y" in points.columns:
        points["index_score"] = points["y"]
    if "metric_value" not in points.columns or "index_score" not in points.columns:
        return _empty_reference_curve_points()
    if "point_order" not in points.columns:
        points["point_order"] = range(1, len(points) + 1)

    original_order = np.arange(1, len(points) + 1)
    point_order = pd.to_numeric(points["point_order"], errors="coerce").to_numpy(dtype="float64")
    point_order = np.where(np.isnan(point_order), original_order, point_order)

    out = pd.DataFrame(
        {
            "point_order": point_order,
            "metric_value": pd.to_numeric(points["metric_value"], errors="coerce").to_numpy(
                dtype="float64"
            ),
            "index_score": pd.to_numeric(points["index_score"], errors="coerce").to_numpy(
                dtype="float64"
            ),
            "original_order": original_order,
        }
    )
    out = out[~(out["metric_value"].isna() & out["index_score"].isna())]
    out = out.sort_values(["point_order", "original_order"], kind="stable")
    out["point_order"] = np.arange(1, len(out) + 1)
    return out[["point_order", "metric_value", "index_score"]].reset_index(drop=True)


def _reference_curve_points_from_row(curve_row: pd.DataFrame | None) -> pd.DataFrame:
    if curve_row is None or len(curve_row) == 0:
        return _empty_reference_curve_points()
    row = curve_row.iloc[[0]]

    if "curve_points" in row.columns:
        stored = row["curve_points"].iloc[0]
        # NOTE(parity): R stores NULL for a missing nested table; the pandas
        # analog is None (or NaN once it has passed through indexing).
        if stored is not None and not (np.isscalar(stored) and _is_na_scalar(stored)):
            normalized = _normalize_reference_curve_points(stored)
            if len(normalized) >= 2:
                return normalized

    point_x_cols = [c for c in row.columns if re.fullmatch(r"curve_point[0-9]+_x", c)]
    if not point_x_cols:
        return _empty_reference_curve_points()
    point_ids = sorted(int(re.sub(r"^curve_point([0-9]+)_x$", r"\1", c)) for c in point_x_cols)

    records = []
    for idx in point_ids:
        x_col, y_col = f"curve_point{idx}_x", f"curve_point{idx}_y"
        if x_col not in row.columns or y_col not in row.columns:
            continue
        x_val = pd.to_numeric(pd.Series([row[x_col].iloc[0]]), errors="coerce").iloc[0]
        y_val = pd.to_numeric(pd.Series([row[y_col].iloc[0]]), errors="coerce").iloc[0]
        if pd.isna(x_val) or pd.isna(y_val):
            continue
        records.append(
            {"point_order": idx, "metric_value": float(x_val), "index_score": float(y_val)}
        )
    return _normalize_reference_curve_points(pd.DataFrame(records))


# =========================================================================== #
# R/14_oh_list_of_metrics.R — SQT List of Metrics xlsx builder
# =========================================================================== #


def oh_list_of_metrics_defaults() -> dict:
    return {
        "data_start_row": 3,
        "data_end_row": 54,
        "col": {
            "functional_category": 1,  # A
            "parameter": 2,  # B
            "metric_units": 3,  # C
            "strat_type": 4,  # D
            "strat_desc": 5,  # E
            "nf_min": 6,  # F
            "nf_max": 7,  # G
            "far_min": 8,  # H
            "far_max": 9,  # I
            "f_min": 10,  # J
            "f_max": 11,  # K
            "applicability": 12,  # L
            "notes": 13,  # M
        },
    }


def format_threshold_cell(value, side="min", higher_is_better=True, boundary=False) -> str:
    if side not in ("min", "max"):
        raise ValueError('side must be one of "min", "max"')
    # R: is.null(value) || length(value) == 0 || !is.finite(value) -> "-"
    # (is.finite() on a character vector is FALSE, so strings also yield "-")
    if value is None or _is_na_scalar(value):
        return "-"
    if not isinstance(value, (bool, int, float, np.bool_, np.integer, np.floating)):
        return "-"
    v = float(value)
    if not math.isfinite(v):
        return "-"

    num_str = _format_number(_signif(v, 3))
    if not boundary:
        return num_str

    # NOTE(parity): the R code computes the same prefix for both
    # higher_is_better branches ("≥ " for min, "≤ " for max).
    prefix = "≥ " if side == "min" else "≤ "
    return prefix + num_str


def oh_threshold_rows_for_metric(threshold_rows, metric_key) -> pd.DataFrame:
    if threshold_rows is None or len(threshold_rows) == 0:
        return pd.DataFrame()
    rows = threshold_rows if isinstance(threshold_rows, pd.DataFrame) else pd.DataFrame(threshold_rows)
    if "metric" not in rows.columns:
        return pd.DataFrame()
    return rows[rows["metric"] == metric_key]


def oh_list_of_metrics_row_values(metric_key, threshold_row, metric_config=None) -> dict:
    entry = oh_metric_entry(metric_key)
    if entry is not None and entry.get("display_name") is not None:
        display_name = entry["display_name"]
    elif (
        metric_config is not None
        and (metric_config.get(metric_key) or {}).get("display_name") is not None
    ):
        display_name = metric_config[metric_key]["display_name"]
    else:
        display_name = metric_key
    units_disp = oh_units_display(metric_key, metric_config)
    if units_disp is not None and str(units_disp) != "":
        metric_label = f"{display_name} ({units_disp})"
    else:
        metric_label = display_name

    higher_is_better = _is_true(_row_scalar(threshold_row, "higher_is_better")) or _is_true(
        ((metric_config or {}).get(metric_key) or {}).get("higher_is_better")
    )

    stratum_label = _row_scalar(threshold_row, "stratum")
    strat_type = ""
    strat_desc = ""
    if _has_stratum_label(stratum_label):
        parts = re.split(r"\s*=\s*|\s*:\s*", str(stratum_label))
        if len(parts) >= 2:
            strat_type = parts[0]
            strat_desc = " ".join(parts[1:])
        else:
            strat_desc = str(stratum_label)

    return {
        "metric_label": metric_label,
        "strat_type": strat_type,
        "strat_desc": strat_desc,
        "nf_min": format_threshold_cell(
            _row_scalar(threshold_row, "not_functioning_min"), "min", higher_is_better
        ),
        "nf_max": format_threshold_cell(
            _row_scalar(threshold_row, "not_functioning_max"), "max", higher_is_better
        ),
        "far_min": format_threshold_cell(
            _row_scalar(threshold_row, "at_risk_min"), "min", higher_is_better
        ),
        "far_max": format_threshold_cell(
            _row_scalar(threshold_row, "at_risk_max"), "max", higher_is_better
        ),
        "f_min": format_threshold_cell(
            _row_scalar(threshold_row, "functioning_min"), "min", higher_is_better
        ),
        "f_max": format_threshold_cell(
            _row_scalar(threshold_row, "functioning_max"), "max", higher_is_better
        ),
        "applicability": "",
        "notes": oh_reference_notes(metric_key),
    }


def _unmerge_intersecting(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Unmerge every merged range that intersects the given rectangle.

    openpyxl raises when writing a non-anchor MergedCell; openxlsx silently
    populated the underlying cells. See module docstring.
    """
    to_unmerge = [
        rng
        for rng in list(ws.merged_cells.ranges)
        if not (
            rng.max_row < min_row
            or rng.min_row > max_row
            or rng.max_col < min_col
            or rng.min_col > max_col
        )
    ]
    for rng in to_unmerge:
        ws.unmerge_cells(str(rng))


def oh_clear_performance_standards_data(ws, opts) -> None:
    """Write empty strings into every data cell rows 3..54 cols A..M — keeps
    the template cell styles intact (R: writeData(NA, keepNA = FALSE))."""
    _unmerge_intersecting(ws, opts["data_start_row"], opts["data_end_row"], 1, opts["col"]["notes"])
    for r in range(opts["data_start_row"], opts["data_end_row"] + 1):
        for c in range(1, opts["col"]["notes"] + 1):
            ws.cell(row=r, column=c).value = ""


def oh_write_performance_standards(ws, context, opts, metric_config=None, mapping=None) -> None:
    threshold_rows = _or((context or {}).get("threshold_rows"), pd.DataFrame())
    category_order = resolved_category_order(mapping)
    current_row = opts["data_start_row"]
    if isinstance(threshold_rows, pd.DataFrame) and "metric" in threshold_rows.columns:
        # NOTE(parity): R keeps NA metric keys here (and would then error in
        # the YAML lookup); NA keys are dropped instead.
        available_metric_keys = threshold_rows["metric"].dropna().unique().tolist()
    else:
        available_metric_keys = []

    for category in category_order:
        metric_keys = metrics_for_resolved_discipline(
            category, mapping=mapping, metric_keys=available_metric_keys
        )
        metric_keys = [
            m for m in metric_keys if len(oh_threshold_rows_for_metric(threshold_rows, m)) > 0
        ]
        if not metric_keys:
            continue

        for metric_key in metric_keys:
            rows_for_metric = oh_threshold_rows_for_metric(threshold_rows, metric_key)
            if len(rows_for_metric) == 0:
                continue
            parameter_name = resolve_metric_function(metric_key, mapping)
            first_row_for_metric = True

            for i in range(len(rows_for_metric)):
                if current_row > opts["data_end_row"]:
                    logger.warning(
                        "List of Metrics: exceeded template data rows; truncating at row %s.",
                        opts["data_end_row"],
                    )
                    return
                values = oh_list_of_metrics_row_values(
                    metric_key,
                    threshold_row=rows_for_metric.iloc[[i]],
                    metric_config=metric_config,
                )

                # Write the functional category and function-based parameter
                # name on every row — the MN template's downward-merged cells
                # are specific to its original layout, so each cell is
                # populated explicitly instead of relying on merges.
                col = opts["col"]
                row_values: dict[int, object] = {
                    col["functional_category"]: category,
                    col["parameter"]: parameter_name,
                    col["metric_units"]: values["metric_label"] if first_row_for_metric else "",
                    col["strat_type"]: values["strat_type"],
                    col["strat_desc"]: values["strat_desc"],
                    col["nf_min"]: values["nf_min"],
                    col["nf_max"]: values["nf_max"],
                    col["far_min"]: values["far_min"],
                    col["far_max"]: values["far_max"],
                    col["f_min"]: values["f_min"],
                    col["f_max"]: values["f_max"],
                    col["applicability"]: values["applicability"],
                    col["notes"]: values["notes"] if first_row_for_metric else "",
                }
                for j in range(1, col["notes"] + 1):
                    ws.cell(row=current_row, column=j).value = _chr(row_values.get(j))

                current_row += 1
                first_row_for_metric = False


def oh_write_references_sheet(ws, metric_config=None, mapping=None) -> None:
    data_start_row = 2
    data_end_row = 28
    n_cols = 4
    _unmerge_intersecting(ws, data_start_row, data_end_row, 1, n_cols)
    for r in range(data_start_row, data_end_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).value = ""

    category_order = resolved_category_order(mapping)
    current_row = data_start_row
    if _mapping_is_usable(mapping):
        metric_keys_covered = mapping["metric_key"].dropna().unique().tolist()
    else:
        metric_keys_covered = oh_covered_metrics()

    for category in category_order:
        metric_keys = metrics_for_resolved_discipline(
            category, mapping=mapping, metric_keys=metric_keys_covered
        )
        if not metric_keys:
            continue
        first_metric_in_cat = True
        for metric_key in metric_keys:
            if current_row > data_end_row:
                break
            display = ((metric_config or {}).get(metric_key) or {}).get("display_name")
            metric_display = display if display is not None else metric_key
            sources = oh_data_sources(metric_key)
            data_source = "; ".join(str(s) for s in sources) if sources else ""

            row_vec = [
                category if first_metric_in_cat else "",
                _chr(resolve_metric_function(metric_key, mapping)),
                metric_display,
                data_source,
            ]
            for j, v in enumerate(row_vec, start=1):
                ws.cell(row=current_row, column=j).value = _chr(v)
            current_row += 1
            first_metric_in_cat = False


def build_oh_list_of_metrics(
    context, template_path, output_path, metric_config=None, mapping=None
):
    """Populate the MN List of Metrics template with project thresholds.

    ``mapping`` is R's ``discipline_function_mapping`` DataFrame
    (metric_key, discipline, function_label, sort_order); None falls back to
    the YAML parameter map.
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise ValueError(f"List of Metrics template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    sheets = wb.sheetnames
    ps_matches = [s for s in sheets if s.lower() == "performance standards"]
    ref_matches = [s for s in sheets if s.lower() == "references"]
    if not ps_matches:
        # R fails here too (subscript out of bounds), just less legibly.
        raise ValueError(
            f"List of Metrics template has no 'Performance Standards' sheet: {template_path}"
        )
    ps_sheet = ps_matches[0]
    # NOTE(parity): R would also error when 'References' is missing (the
    # is.na() guard is unreachable); the port skips instead, honoring the
    # guard's evident intent.
    ref_sheet = ref_matches[0] if ref_matches else None

    opts = oh_list_of_metrics_defaults()
    oh_clear_performance_standards_data(wb[ps_sheet], opts)
    oh_write_performance_standards(
        wb[ps_sheet], context, opts, metric_config=metric_config, mapping=mapping
    )
    if ref_sheet is not None:
        oh_write_references_sheet(wb[ref_sheet], metric_config=metric_config, mapping=mapping)

    wb.save(output_path)
    return output_path


# =========================================================================== #
# R/15_oh_sqt_workbook.R — SQT Reference Curves xlsx builder
# =========================================================================== #

# --------------------------------------------------------------------------- #
# Data extraction helpers
# --------------------------------------------------------------------------- #


def oh_sqt_ohio_points_for_metric(threshold_rows, metric_key) -> list[float]:
    """Curve breakpoint x-coordinates for the metric, sorted by index score
    ascending."""
    if threshold_rows is None:
        return []
    rows = threshold_rows if isinstance(threshold_rows, pd.DataFrame) else pd.DataFrame(threshold_rows)
    if "metric" not in rows.columns:
        return []
    rows = rows[rows["metric"] == metric_key]
    if len(rows) == 0:
        return []

    row = rows.iloc[[0]]

    points = _reference_curve_points_from_row(row)
    if len(points) > 0:
        ordered = points.sort_values("index_score", kind="stable")
        return [float(v) for v in ordered["metric_value"].tolist()]

    # Flat-column fallback kept from R (curve_point1..3 only, finite pairs).
    xs, ys = [], []
    for idx in (1, 2, 3):
        x = _row_scalar(row, f"curve_point{idx}_x")
        y = _row_scalar(row, f"curve_point{idx}_y")
        x = float(x) if isinstance(x, (int, float, np.integer, np.floating)) else float("nan")
        y = float(y) if isinstance(y, (int, float, np.integer, np.floating)) else float("nan")
        xs.append(x)
        ys.append(y)
    keep = [math.isfinite(x) and math.isfinite(y) for x, y in zip(xs, ys)]
    xs = [x for x, k in zip(xs, keep) if k]
    ys = [y for y, k in zip(ys, keep) if k]
    if not xs:
        return []
    order = np.argsort(np.asarray(ys), kind="stable")
    return [xs[i] for i in order]


def oh_sqt_threshold_rows_for_metric(threshold_rows, metric_key) -> pd.DataFrame:
    # Identical to oh_threshold_rows_for_metric (duplicated in the R sources).
    return oh_threshold_rows_for_metric(threshold_rows, metric_key)


def oh_sqt_index_values(n_points: int, higher_is_better: bool = True) -> list[float]:
    """Standard index-value sequence for N breakpoints, optionally mirrored
    for "higher is worse" metrics. Based on the WI go-by convention."""
    if n_points <= 0:
        return []
    if n_points == 2:
        iv = [0.0, 1.0]
    elif n_points == 3:
        iv = [0.0, 0.70, 1.00]
    elif n_points == 4:
        iv = [0.0, 0.30, 0.70, 1.00]
    else:
        iv = [float(v) for v in np.linspace(0.0, 1.0, n_points)]
    if not _is_true(higher_is_better):  # R: if (!isTRUE(higher_is_better)) rev(iv)
        iv = iv[::-1]
    return iv


def oh_sqt_linear_segments(points, iv) -> dict:
    """Per-segment linear coefficients (Y = a*X + b) between consecutive
    breakpoints. Returns {"seg_labels": [...], "slopes": [...],
    "intercepts": [...]} (R named-list keys)."""
    points = [float(p) for p in points]
    iv = [float(v) for v in iv]
    n_seg = len(points) - 1
    if n_seg < 1:
        return {"seg_labels": [], "slopes": [], "intercepts": []}
    slopes: list[float] = []
    intercepts: list[float] = []
    seg_labels: list[str] = []
    for j in range(n_seg):
        x1, x2 = points[j], points[j + 1]
        y1, y2 = iv[j], iv[j + 1]
        if math.isfinite(x1) and math.isfinite(x2) and (x2 - x1) != 0:
            a = (y2 - y1) / (x2 - x1)
            slopes.append(a)
            intercepts.append(y1 - a * x1)
        else:
            slopes.append(float("nan"))
            intercepts.append(float("nan"))
        seg_labels.append(
            f"{_format_number(_signif(y1, 3))} → {_format_number(_signif(y2, 3))}"
        )
    return {"seg_labels": seg_labels, "slopes": slopes, "intercepts": intercepts}


def oh_sqt_band(y) -> str | None:
    """Band label (NF/FAR/F) for a single index value; None when non-finite."""
    if not isinstance(y, (bool, int, float, np.bool_, np.integer, np.floating)) or not math.isfinite(
        float(y)
    ):
        return None
    y = float(y)
    if y < 0.30:
        return "NF"
    if y < 0.70:
        return "FAR"
    return "F"


def oh_sqt_segment_band_labels(iv) -> list[str]:
    """Per-segment performance-band label. For segment i spanning
    iv[i] -> iv[i+1], the bands lying in the open interval (y_lo, y_hi),
    joined by "/" (em dash when none)."""
    iv = [float(v) for v in iv]
    n = len(iv) - 1
    if n < 1:
        return []
    lbls = []
    for i in range(n):
        y_lo = min(iv[i], iv[i + 1])
        y_hi = max(iv[i], iv[i + 1])
        bands = []
        if y_lo < 0.30 and y_hi > 0.00:
            bands.append("NF")
        if y_lo < 0.70 and y_hi > 0.30:
            bands.append("FAR")
        if y_hi > 0.70:
            bands.append("F")
        lbls.append("/".join(bands) if bands else "—")
    return lbls


# --------------------------------------------------------------------------- #
# Styles
# --------------------------------------------------------------------------- #


@dataclass(frozen=True)
class CellStyle:
    """openpyxl analog of an openxlsx createStyle() bundle."""

    font: Font | None = None
    fill: PatternFill | None = None
    border: Border | None = None
    alignment: Alignment | None = None
    number_format: str | None = None


def _solid_fill(hex6: str) -> PatternFill:
    argb = "FF" + hex6.lstrip("#").upper()
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def _border_all(hex6: str = "7F7F7F") -> Border:
    side = Side(style="thin", color="FF" + hex6.lstrip("#").upper())
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_style(cell, style: CellStyle) -> None:
    if style.font is not None:
        cell.font = style.font
    if style.fill is not None:
        cell.fill = style.fill
    if style.border is not None:
        cell.border = style.border
    if style.alignment is not None:
        cell.alignment = style.alignment
    if style.number_format is not None:
        cell.number_format = style.number_format


def _style_range(ws, style: CellStyle, rows, cols) -> None:
    """openxlsx addStyle(..., gridExpand = TRUE): style every cell in the
    rows x cols rectangle (styles on merged non-anchor cells are legal)."""
    for r in rows:
        for c in cols:
            _apply_style(ws.cell(row=r, column=c), style)


def oh_sqt_build_styles() -> dict[str, CellStyle]:
    def disc_style(fill_hex: str) -> CellStyle:
        return CellStyle(
            font=Font(size=13, bold=True),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=_solid_fill(fill_hex),
            border=_border_all("7F7F7F"),
        )

    return {
        "title": CellStyle(
            font=Font(size=16, bold=True, color="FFFFFFFF"),
            alignment=Alignment(horizontal="left", vertical="center"),
            fill=_solid_fill("1F4E79"),
        ),
        "subtitle": CellStyle(
            font=Font(size=10, italic=True, color="FF595959"),
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        ),
        "metric_label": CellStyle(
            font=Font(size=12, bold=True),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=_solid_fill("D9E1F2"),
            border=_border_all("7F7F7F"),
        ),
        "function_label": CellStyle(
            font=Font(size=10, italic=True, color="FF595959"),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        "stmt": CellStyle(
            font=Font(size=10, italic=True, color="FF595959"),
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        ),
        "row_label": CellStyle(
            font=Font(bold=True),
            fill=_solid_fill("F2F2F2"),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "seg_header": CellStyle(
            font=Font(size=10, bold=True, color="FF404040"),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        "coef_header": CellStyle(
            font=Font(size=10, italic=True, color="FF404040"),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "num": CellStyle(
            alignment=Alignment(horizontal="center"),
            number_format="0.####",
        ),
        "idx_red": CellStyle(
            font=Font(size=11, bold=True),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=_solid_fill("F8CBAD"),
            number_format="0.##",
            border=_border_all("7F7F7F"),
        ),
        "idx_yellow": CellStyle(
            font=Font(size=11, bold=True),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=_solid_fill("FFE699"),
            number_format="0.##",
            border=_border_all("7F7F7F"),
        ),
        "idx_green": CellStyle(
            font=Font(size=11, bold=True),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=_solid_fill("C6E0B4"),
            number_format="0.##",
            border=_border_all("7F7F7F"),
        ),
        "discipline_hydrology": disc_style("DCE6F1"),
        "discipline_hydraulics": disc_style("B4C7E7"),
        "discipline_geomorphology": disc_style("FBE5D6"),
        "discipline_physicochemistry": disc_style("FFF2CC"),
        "discipline_biology": disc_style("E2EFDA"),
    }


def oh_sqt_band_fill_style(styles: dict, y) -> CellStyle:
    """Pick the index-band fill style for a numeric index value."""
    if not isinstance(y, (bool, int, float, np.bool_, np.integer, np.floating)) or not math.isfinite(
        float(y)
    ):
        return styles["num"]
    y = float(y)
    if y < 0.30:
        return styles["idx_red"]
    if y < 0.70:
        return styles["idx_yellow"]
    return styles["idx_green"]


def oh_sqt_discipline_style(styles: dict, discipline: str) -> CellStyle:
    key = "discipline_" + re.sub(r"[^A-Za-z]", "", str(discipline)).lower()
    return styles.get(key, styles["metric_label"])


# --------------------------------------------------------------------------- #
# Horizontal layout plan
# --------------------------------------------------------------------------- #


def oh_sqt_layout_plan(
    metrics_by_discipline: dict,
    block_height: int = 22,
    block_gap: int = 2,
    band_cols: int = 8,
    header_row: int = 5,
    body_start_row: int = 8,
) -> dict:
    """Per-discipline column ranges and per-metric (row_start, row_end)
    offsets. Disciplines placed left-to-right; col A (1) is a gutter and
    discipline i starts at 2 + (i-1) * band_cols."""
    out: dict = {}
    col_start = 2
    for disc, metrics in metrics_by_discipline.items():
        metrics = _or(metrics, [])
        blocks: dict = {}
        current_row = body_start_row
        for mk in metrics:
            blocks[mk] = {
                "metric_key": mk,
                "row_start": current_row,
                "row_end": current_row + block_height - 1,
            }
            current_row += block_height + block_gap
        out[disc] = {
            "name": disc,
            "col_start": col_start,
            "col_end": col_start + band_cols - 1,
            "header_row": header_row,
            "blocks": blocks,
        }
        col_start += band_cols
    return out


# --------------------------------------------------------------------------- #
# Writers
# --------------------------------------------------------------------------- #


def oh_sqt_write_discipline_header_row(ws, layout: dict, styles: dict) -> None:
    for disc, band in layout.items():
        ws.cell(row=band["header_row"], column=band["col_start"], value=str(disc).upper())
        ws.merge_cells(
            start_row=band["header_row"],
            start_column=band["col_start"],
            end_row=band["header_row"],
            end_column=band["col_end"],
        )
        _style_range(
            ws,
            oh_sqt_discipline_style(styles, disc),
            [band["header_row"]],
            range(band["col_start"], band["col_end"] + 1),
        )
    if layout:
        first_band = next(iter(layout.values()))
        ws.row_dimensions[first_band["header_row"]].height = 24


def oh_sqt_write_metric_block(
    ws,
    metric_key,
    context,
    metric_config,
    bundle_dir,
    col_start: int,
    row_start: int,
    styles: dict,
    function_label=None,
    stratum_idx: int = 0,
    block_height: int = 22,
) -> int:
    """Write a single metric block (8 cols x ~22 rows) at absolute sheet
    coordinates. ``stratum_idx`` is 0-based here (R used 1-based). Returns
    the last row of the block footprint."""
    context = _or(context, {})
    mc = ((metric_config or {}).get(metric_key)) or {}
    display = _or(mc.get("display_name"), metric_key)
    units = oh_units_display(metric_key, metric_config)
    if units is not None and str(units) != "":
        label = f"{display} ({units})"
    else:
        label = display

    rows_for_metric = oh_sqt_threshold_rows_for_metric(context.get("threshold_rows"), metric_key)
    if stratum_idx < len(rows_for_metric):
        stratum_row = rows_for_metric.iloc[[stratum_idx]]
    else:
        stratum_row = rows_for_metric.iloc[0:0]
    higher_is_better = _is_true(_row_scalar(stratum_row, "higher_is_better"))
    stratum_label = _row_scalar(stratum_row, "stratum")
    has_stratum = _has_stratum_label(stratum_label)

    header_text = f"{label} — {stratum_label}" if has_stratum else label

    col_end = col_start + 7  # 8-col discipline band
    data_col_start = col_start + 1  # first band column is the row label

    # Row 0: metric header (merged)
    hdr_row = row_start
    ws.cell(row=hdr_row, column=col_start, value=header_text)
    ws.merge_cells(start_row=hdr_row, start_column=col_start, end_row=hdr_row, end_column=col_end)
    _style_range(ws, styles["metric_label"], [hdr_row], range(col_start, col_end + 1))
    ws.row_dimensions[hdr_row].height = 22

    # Row 1: function label sub-header (optional)
    fn_row = row_start + 1
    if not _is_na_scalar(function_label) and str(function_label) != "":
        ws.cell(row=fn_row, column=col_start, value=str(function_label))
        ws.merge_cells(start_row=fn_row, start_column=col_start, end_row=fn_row, end_column=col_end)
        _style_range(ws, styles["function_label"], [fn_row], range(col_start, col_end + 1))

    points = oh_sqt_ohio_points_for_metric(stratum_row, metric_key)
    iv = oh_sqt_index_values(len(points), higher_is_better)

    # Row 2: Field Value
    fv_row = row_start + 2
    ws.cell(row=fv_row, column=col_start, value="Field Value")
    _apply_style(ws.cell(row=fv_row, column=col_start), styles["row_label"])
    if points:
        for k, x in enumerate(points):
            cell = ws.cell(row=fv_row, column=data_col_start + k)
            cell.value = float(x)
            _apply_style(cell, styles["num"])
    else:
        ws.cell(row=fv_row, column=data_col_start, value="No curve data available.")

    # Row 3: Index Value with red/yellow/green per-cell banding
    iv_row = row_start + 3
    ws.cell(row=iv_row, column=col_start, value="Index Value")
    _apply_style(ws.cell(row=iv_row, column=col_start), styles["row_label"])
    for k, y in enumerate(iv):
        cell = ws.cell(row=iv_row, column=data_col_start + k)
        cell.value = float(y)
        _apply_style(cell, oh_sqt_band_fill_style(styles, y))

    # Row 4: blank spacer. Row 5: NF/FAR/F segment headers.
    seg_row = row_start + 5
    if len(iv) >= 2:
        seg_bands = oh_sqt_segment_band_labels(iv)
        for k, lbl in enumerate(seg_bands):
            cell = ws.cell(row=seg_row, column=data_col_start + k)
            cell.value = lbl
            _apply_style(cell, styles["seg_header"])

    # Row 6: Coefficients label (merged)
    coef_header_row = row_start + 6
    ws.cell(row=coef_header_row, column=col_start, value="Coefficients — Y = a * X + b")
    ws.merge_cells(
        start_row=coef_header_row,
        start_column=col_start,
        end_row=coef_header_row,
        end_column=col_end,
    )
    _style_range(ws, styles["coef_header"], [coef_header_row], range(col_start, col_end + 1))

    # Rows 7-8: a (slope) and b (intercept)
    a_row = row_start + 7
    b_row = row_start + 8
    ws.cell(row=a_row, column=col_start, value="a")
    ws.cell(row=b_row, column=col_start, value="b")
    _style_range(ws, styles["row_label"], [a_row, b_row], [col_start])
    if len(points) >= 2:
        segs = oh_sqt_linear_segments(points, iv)
        n_seg = len(segs["slopes"])
        for k in range(n_seg):
            a_val = segs["slopes"][k]
            b_val = segs["intercepts"][k]
            # openxlsx keepNA=FALSE writes NA as blank; openpyxl can't store NaN.
            ws.cell(row=a_row, column=data_col_start + k).value = (
                float(a_val) if math.isfinite(a_val) else None
            )
            ws.cell(row=b_row, column=data_col_start + k).value = (
                float(b_val) if math.isfinite(b_val) else None
            )
        _style_range(
            ws, styles["num"], [a_row, b_row], range(data_col_start, data_col_start + n_seg)
        )

    # Rows 10+: embedded plot sized to fit within the 8-col band.
    # R used context$metrics[[metric_key]]$plot_file under bundle_dir; the
    # port also accepts "plot_png" (raw PNG bytes or zero-arg callable
    # returning them). Skipped when absent — see module docstring.
    entry = ((context.get("metrics") or {}).get(metric_key)) or {}
    plot_start = row_start + 10
    img = None
    plot_rel = entry.get("plot_file")
    if plot_rel is not None and not _is_na_scalar(plot_rel) and str(plot_rel) != "" and bundle_dir is not None:
        plot_path = Path(bundle_dir) / str(plot_rel)
        if plot_path.exists():
            img = XLImage(str(plot_path))
    if img is None:
        png = entry.get("plot_png")
        if callable(png):
            png = png()
        if png:
            img = XLImage(BytesIO(png))
    if img is not None:
        img.width = 346  # R: 3.6 in x 2.3 in @ 96 dpi
        img.height = 221
        ws.add_image(img, f"{get_column_letter(col_start)}{plot_start}")

    return row_start + block_height - 1


# --------------------------------------------------------------------------- #
# Main entry point
# --------------------------------------------------------------------------- #


def build_oh_reference_curves_workbook(
    context,
    output_path,
    metric_config=None,
    strat_config=None,
    bundle_dir=None,
    mapping=None,
):
    """Build the SQT Reference Curves workbook from scratch.

    ``mapping`` is R's ``discipline_function_mapping`` DataFrame; R's unused
    ``template_path`` argument is dropped (see module docstring).
    """
    context = _or(context, {})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reference Standards"

    # Column widths: col A gutter; then 5 x 8 band-columns for labels + data.
    ws.column_dimensions["A"].width = 2
    for i in range(5):
        base = 2 + i * 8
        ws.column_dimensions[get_column_letter(base)].width = 16
        for c in range(base + 1, base + 7):
            ws.column_dimensions[get_column_letter(c)].width = 9
        ws.column_dimensions[get_column_letter(base + 7)].width = 3

    styles = oh_sqt_build_styles()

    # Title + subtitle
    ws.cell(row=1, column=2, value="Stream Quantification Tool — Reference Curves")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=41)
    _style_range(ws, styles["title"], [1], range(2, 42))
    ws.row_dimensions[1].height = 28

    session_meta = _or(context.get("session_meta"), {})
    # NOTE(parity): R falls back to format(Sys.time(), "%Y-%m-%d %H:%M") when
    # the caller supplies no generated_at; kept despite the no-nondeterminism
    # rule because it is R behavior (callers should always pass it).
    generated_at = _or(session_meta.get("generated_at"), datetime.now().strftime("%Y-%m-%d %H:%M"))
    complete_metrics = _or(session_meta.get("complete_metrics"), 0)
    subtitle_text = (
        f"Session: {generated_at}"
        f"  |  Complete metrics: {complete_metrics}"
        f"  |  Performance bands: NF < 0.30 ≤ FAR < 0.70 ≤ F"
    )
    ws.cell(row=2, column=2, value=subtitle_text)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=41)
    _style_range(ws, styles["subtitle"], [2], range(2, 42))

    # metrics_by_discipline via the resolver (mapping wins, YAML fallback).
    threshold_rows = context.get("threshold_rows")
    if isinstance(threshold_rows, pd.DataFrame) and "metric" in threshold_rows.columns:
        available_metric_keys = threshold_rows["metric"].dropna().unique().tolist()
    else:
        available_metric_keys = []
    disc_order = resolved_category_order(mapping)
    if not disc_order:
        disc_order = fixed_discipline_order()

    metrics_by_discipline = {
        disc: metrics_for_resolved_discipline(
            disc, mapping=mapping, metric_keys=available_metric_keys
        )
        for disc in disc_order
    }

    layout = oh_sqt_layout_plan(metrics_by_discipline)
    oh_sqt_write_discipline_header_row(ws, layout, styles)

    # Rows 6-7 between the discipline header and the first metric blocks are
    # intentional visual padding.
    for disc, band in layout.items():
        current_row = band["header_row"] + 3  # first block starts at row 8
        for mk in band["blocks"]:
            rows_for_metric = oh_sqt_threshold_rows_for_metric(threshold_rows, mk)
            strata_count = max(1, len(rows_for_metric))
            function_label = resolve_metric_function(mk, mapping)
            for s in range(strata_count):
                last_row = oh_sqt_write_metric_block(
                    ws,
                    mk,
                    context,
                    metric_config=metric_config,
                    bundle_dir=bundle_dir,
                    col_start=band["col_start"],
                    row_start=current_row,
                    styles=styles,
                    function_label=function_label,
                    stratum_idx=s,
                )
                # 22-row block + 2-row gap (+1 already in last_row index)
                current_row = last_row + 2 + 1

    # Pull Down Notes sheet (stream-type validation values)
    pdn = wb.create_sheet("Pull Down Notes")
    pdn.column_dimensions["A"].width = 14
    pdn.column_dimensions["B"].width = 18
    pdn.cell(row=1, column=1, value="Stream Type:")
    _apply_style(pdn.cell(row=1, column=1), styles["row_label"])
    levels = ((strat_config or {}).get("StreamType2") or {}).get("levels")
    stream_types = list(_or(levels, ["B", "C", "E", "F"]))
    for i, stream_type in enumerate(stream_types):
        pdn.cell(row=2 + i, column=2, value=stream_type)

    wb.save(output_path)
    return output_path
