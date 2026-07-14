"""Export workspace — port of app/modules/mod_summary_export.R.

The read-only export dashboard shown in the workspace modal (opened from the
Reference Curves summary page's "Export" button). Renders session-status cards
and four download handlers: the OH List of Metrics workbook, the SQT Reference
Curves workbook, the Science Support Document (HTML), and the DEEP assessment
bundle (JSON). Exports gate on the discipline -> function mapping being
confirmed, valid, and covering every summary-eligible metric.
"""

from __future__ import annotations

import json
import logging
import os
import tempfile
from datetime import datetime

from shiny import module, reactive, render, req, ui

from streamcurves.deep_export import (
    deep_collect_curve_rows,
    deep_slug,
    write_deep_assessment_bundle,
)
from streamcurves.oh_export import (
    build_oh_list_of_metrics,
    build_oh_reference_curves_workbook,
)
from streamcurves.paths import TEMPLATES_DIR
from streamcurves.science_report import build_science_support_html
from views import assessment_publish as ap
from views import summary_state as sst
from views.state import AppState
from views.theme import bi
from views.uihelpers import explanation_card

logger = logging.getLogger("streamcurves")

_LIST_OF_METRICS_TEMPLATE = TEMPLATES_DIR / "MN-List-of-Metricsv2.0.xlsx"
_SQT_TEMPLATE = TEMPLATES_DIR / "WISQT_Reference_Curves.xlsx"  # noqa: F841 (built from scratch)


# --------------------------------------------------------------------------- #
# Error-artifact fallbacks (R write_error_xlsx / write_error_html).
# --------------------------------------------------------------------------- #


def _error_xlsx_bytes(title: str, message: str) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Error"
    ws["A1"] = title
    ws["A2"] = str(message)
    buf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb.save(buf.name)
        buf.close()
        with open(buf.name, "rb") as f:
            return f.read()
    finally:
        try:
            os.unlink(buf.name)
        except OSError:
            pass


def _error_html_bytes(title: str, message: str) -> bytes:
    import html as _html

    return (
        f"<!DOCTYPE html><html><head><meta charset='utf-8'><title>{_html.escape(title)}"
        f"</title></head><body><h1>{_html.escape(title)}</h1>"
        f"<pre>{_html.escape(str(message))}</pre></body></html>"
    ).encode("utf-8")


def _xlsx_to_bytes(build) -> bytes:
    """Run an openpyxl builder that writes to output_path -> return the bytes."""
    with tempfile.TemporaryDirectory() as d:
        out = os.path.join(d, "out.xlsx")
        build(out)
        with open(out, "rb") as f:
            return f.read()


@module.ui
def summary_export_ui():
    return ui.output_ui("summary_page")


@module.server
def summary_export_server(input, output, session, state: AppState):
    ns = session.ns

    @reactive.calc
    def export_context():
        req(state.data() is not None)
        ctx = sst.build_summary_export_context(state)
        ctx["session_meta"]["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return ctx

    def _cfg():
        with reactive.isolate():
            return {
                "metric_config": state.metric_config() or {},
                "strat_config": state.strat_config() or {},
                "mapping": state.discipline_function_mapping(),
            }

    def _report_enabled(key: str, default: bool = True) -> bool:
        with reactive.isolate():
            output_config = state.output_config() or {}
        products = output_config.get("report_products") or {}
        entry = products.get(key)
        if not isinstance(entry, dict) or "enabled" not in entry:
            return default
        return bool(entry["enabled"])

    # ── Dashboard (R output$summary_page) ─────────────────────────────────────
    # suspend_when_hidden=False is needed so the body renders inside the modal
    # (Shiny treats modal content as hidden), but the export context is
    # expensive (a snapshot per eligible metric) — so gate on the modal actually
    # being open for this type, else it would compute eagerly at data-load and
    # stall the Reference Curves summary page.
    @output(suspend_when_hidden=False)
    @render.ui
    def summary_page():
        req(state.data() is not None)
        req(state.workspace_modal_type() == "summary_export")
        ctx = export_context()
        meta = ctx["session_meta"]
        confirmed = ctx["discipline_function_mapping_confirmed"]
        # The DEEP bundle needs at least one finalized Phase-4 curve, else
        # build_bundle_from_state raises and the download would otherwise be an error
        # stub (not a real .deep.json). Gate the button on the same rule the builder
        # uses so no broken bundle can ever be produced.
        deep_ready = bool(deep_collect_curve_rows(state.completed_metrics() or {}))

        def stat_card(title, value, detail=None):
            body = [
                ui.div(str(value), class_="export-stat-val fw-bold fs-4"),
                ui.div(title, class_="export-stat-lab text-muted small"),
            ]
            if detail:
                body.append(ui.div(detail, class_="text-muted", style="font-size:0.72rem;"))
            return ui.div(ui.div(*body, class_="card-body py-2"), class_="card")

        def download_card(title, description, control):
            return ui.div(
                ui.div(
                    ui.tags.h6(title, class_="fw-bold mb-1"),
                    ui.tags.p(description, class_="text-muted small mb-2"),
                    control,
                    class_="card-body",
                ),
                class_="card mb-2",
            )

        def dl_button(dl_id, label, *, enabled=True, disabled_title=None):
            cls = "btn btn-primary btn-sm"
            if not confirmed:
                return ui.download_button(
                    ns(dl_id),
                    label,
                    class_=cls + " disabled",
                    title="Confirm the Discipline → Function → Metric mapping first.",
                    aria_disabled="true",
                )
            if not enabled:
                return ui.download_button(
                    ns(dl_id),
                    label,
                    class_=cls + " disabled",
                    title=disabled_title or "Not ready to export yet.",
                    aria_disabled="true",
                )
            return ui.download_button(ns(dl_id), label, class_=cls)

        parts = [
            explanation_card(
                "Export SQT Deliverables",
                ui.tags.p(
                    "Download the finalized reference curves as regulatory SQT "
                    "deliverables and a portable assessment bundle for the DEEP executor app.",
                    class_="mb-0",
                ),
            ),
            ui.div(
                ui.card_header("Session Status"),
                ui.div(
                    stat_card("Summary metrics", meta["metric_count"]),
                    stat_card("Curves ready", meta["complete_metrics"]),
                    stat_card("Needs review", meta["review_metrics"]),
                    stat_card("Manual curves", meta["manual_curve_metrics"]),
                    stat_card(
                        "Regional curves",
                        meta["regional_curve_sets"],
                        f"{meta['regional_curve_rows']} rows available",
                    ),
                    class_="d-flex gap-2 flex-wrap p-2",
                ),
                class_="card mb-3",
            ),
        ]

        if not confirmed:
            parts.append(
                ui.div(
                    bi("exclamation-triangle-fill"),
                    " The Discipline → Function → Metric mapping must be reviewed and "
                    "saved (and cover every summary metric) before exports unlock. "
                    "Open the mapping on the Reference Curves page.",
                    class_="alert alert-warning",
                )
            )

        sqt_cards = []
        if _report_enabled("list_of_metrics"):
            sqt_cards.append(
                download_card(
                    "List of Metrics (xlsx)",
                    "The MN-style List of Metrics workbook populated with this "
                    "project's performance-standard thresholds.",
                    dl_button("dl_list_of_metrics", "Download List of Metrics"),
                )
            )
        if _report_enabled("sqt_workbook"):
            sqt_cards.append(
                download_card(
                    "SQT Reference Curves (xlsx)",
                    "The Reference Curves workbook — per-metric discipline bands, "
                    "line coefficients, and embedded curve figures.",
                    dl_button("dl_sqt_workbook", "Download Reference Curves"),
                )
            )
        if _report_enabled("science_support"):
            sqt_cards.append(
                download_card(
                    "Science Support Document (HTML)",
                    "A self-contained narrative of the reference-curve development, "
                    "organized by the Stream Functions Pyramid Framework.",
                    dl_button("dl_science_support", "Download Science Support"),
                )
            )
        if sqt_cards:
            parts.append(
                ui.div(
                    ui.card_header("SQT Deliverables"),
                    ui.div(*sqt_cards, class_="card-body"),
                    class_="card mb-3",
                )
            )

        parts.append(
            ui.div(
                ui.card_header("DEEP Assessment Bundle"),
                ui.div(
                    download_card(
                        "DEEP Assessment Bundle (.deep.json)",
                        "A portable detailed-assessment definition for the DEEP "
                        "executor app — finalized curves, mapped to STAF functions.",
                        dl_button(
                            "dl_deep_assessment",
                            "Download DEEP Bundle",
                            enabled=deep_ready,
                            disabled_title=(
                                "Finalize at least one reference curve (Phase 4) "
                                "before exporting the DEEP bundle."
                            ),
                        ),
                    ),
                    ui.div(
                        bi("magic"),
                        " This bundle is a testable snapshot. To make it a reusable, "
                        "versioned assessment that DEEP lists for everyone, open the ",
                        ui.tags.strong("Publish"),
                        " page (or click Save, top right) and publish this session.",
                        class_="alert alert-info mt-2 mb-0 small",
                    ),
                    class_="card-body",
                ),
                class_="card mb-3",
            )
        )
        return ui.div(*parts)

    # ── Download handlers ─────────────────────────────────────────────────────
    @render.download(filename="List-of-Metrics_v0.1.xlsx")
    def dl_list_of_metrics():
        try:
            ctx = export_context()
            cfg = _cfg()
            data = _xlsx_to_bytes(
                lambda out: build_oh_list_of_metrics(
                    ctx,
                    str(_LIST_OF_METRICS_TEMPLATE),
                    out,
                    metric_config=cfg["metric_config"],
                    mapping=cfg["mapping"],
                )
            )
        except Exception as e:  # noqa: BLE001
            logger.exception("List of Metrics export failed")
            data = _error_xlsx_bytes("List of Metrics export failed", e)
        yield data

    @render.download(filename="SQT_Reference_Curves_v0.1.xlsx")
    def dl_sqt_workbook():
        try:
            ctx = export_context()
            cfg = _cfg()
            data = _xlsx_to_bytes(
                lambda out: build_oh_reference_curves_workbook(
                    ctx,
                    out,
                    metric_config=cfg["metric_config"],
                    strat_config=cfg["strat_config"],
                    mapping=cfg["mapping"],
                )
            )
        except Exception as e:  # noqa: BLE001
            logger.exception("SQT workbook export failed")
            data = _error_xlsx_bytes("SQT Reference Curves export failed", e)
        yield data

    @render.download(filename="science_support.html")
    def dl_science_support():
        try:
            ctx = export_context()
            cfg = _cfg()
            html = build_science_support_html(ctx, cfg["metric_config"])
        except Exception as e:  # noqa: BLE001
            logger.exception("Science Support export failed")
            html = _error_html_bytes("Science Support export failed", e).decode("utf-8")
        yield html.encode("utf-8")

    @render.download(
        filename=lambda: (
            deep_slug(state.isolate_get("session_name") or "spring-assessment")
            + ".deep.json"
        )
    )
    def dl_deep_assessment():
        # Shared builder: finalized curves + mapping + config + region-of-applicability.
        try:
            bundle = ap.build_bundle_from_state(state)
        except ValueError as e:  # no finalized curves yet
            ui.notification_show(str(e), type="warning", duration=8)
            yield json.dumps(
                {
                    "error": str(e),
                    "hint": "Complete at least one metric's Phase 4 curve, then export.",
                },
                indent=2,
            ).encode("utf-8")
            return
        except Exception as e:  # noqa: BLE001
            logger.exception("DEEP export failed")
            yield json.dumps({"error": str(e), "hint": "DEEP export failed."}, indent=2).encode(
                "utf-8"
            )
            return
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "bundle.deep.json")
            write_deep_assessment_bundle(bundle, out)
            with open(out, "rb") as f:
                yield f.read()
