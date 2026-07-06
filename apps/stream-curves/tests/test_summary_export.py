"""Tests for the export workspace error-artifact fallbacks (M6 part 4).

The happy-path exporters (build_oh_list_of_metrics / _reference_curves_workbook /
build_deep_assessment_bundle) are golden-tested in test_oh_export.py and
test_deep_export.py; the full build_summary_export_context -> download flow is
browser-verified. These cover the new error-artifact helpers that keep a failed
export from breaking the download.
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from shiny import reactive

from streamcurves.cleaning import clean_data
from streamcurves.deep_export import deep_collect_curve_rows
from streamcurves.derive import derive_variables
from streamcurves.oh_export import (
    build_oh_list_of_metrics,
    build_oh_reference_curves_workbook,
)
from streamcurves.paths import TEMPLATES_DIR
from streamcurves.precheck import run_metric_precheck
from streamcurves.science_report import build_science_support_html
from streamcurves.staf_library import staf_metric_library_default_mapping
from streamcurves.workbook import read_input_workbook
from tests.golden_io import GOLDEN_DIR
from views import summary_state as ss
from views.state import AppState
from views.summary_export import (
    _error_html_bytes,
    _error_xlsx_bytes,
    _xlsx_to_bytes,
)

FIXTURE = GOLDEN_DIR.parent / "fixtures" / "OSAM_summarydata.xlsx"


def test_error_xlsx_is_a_valid_workbook():
    data = _error_xlsx_bytes("List of Metrics export failed", "boom: bad thing")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws["A1"].value == "List of Metrics export failed"
    assert "boom" in str(ws["A2"].value)


def test_error_html_is_escaped_and_wrapped():
    data = _error_html_bytes("SQT failed", "<script>alert(1)</script>")
    text = data.decode("utf-8")
    assert text.startswith("<!DOCTYPE html>")
    assert "SQT failed" in text
    # the message is HTML-escaped (no raw script tag)
    assert "<script>" not in text
    assert "&lt;script&gt;" in text


def test_xlsx_to_bytes_runs_builder_and_returns_bytes():
    def _build(out_path):
        wb = openpyxl.Workbook()
        wb.active["A1"] = "hello"
        wb.save(out_path)

    data = _xlsx_to_bytes(_build)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.active["A1"].value == "hello"


# --------------------------------------------------------------------------- #
# End-to-end: build_summary_export_context -> all four exporters
# --------------------------------------------------------------------------- #


@pytest.fixture(scope="module")
def export_state() -> AppState:
    """A loaded OSAM state with one finalized (perRiffle) reference curve."""
    state = AppState.fresh()
    bundle = read_input_workbook(FIXTURE)
    cleaned, qa = clean_data(
        bundle["raw_data"], bundle["metric_config"],
        bundle["strat_config"], bundle["factor_recode_config"],
    )
    dat = derive_variables(
        cleaned, bundle["factor_recode_config"],
        bundle["predictor_config"], bundle["strat_config"],
    )
    with reactive.isolate():
        state.metric_config.set(bundle["metric_config"])
        state.strat_config.set(bundle["strat_config"])
        state.predictor_config.set(bundle["predictor_config"])
        state.factor_recode_config.set(bundle["factor_recode_config"])
        state.data.set(dat)
        state.qa_log.set(qa)
        state.precheck_df.set(run_metric_precheck(dat, bundle["metric_config"]))
        state.current_metric.set("perRiffle")
        state.app_data_loaded.set(True)
        # seed + confirm the discipline mapping so the export gate would pass
        mapping = staf_metric_library_default_mapping(
            list(bundle["metric_config"].keys()), bundle["metric_config"]
        )
        state.discipline_function_mapping.set(mapping)

    # finalize a perRiffle curve (phase1 -> phase4 backfill)
    with reactive.isolate():
        ss.set_metric_curve_stratification(state, "perRiffle", "none")
        b1 = ss.build_metric_phase1_backfill(state, "perRiffle", mode="full")
        ss.commit_metric_phase1_backfill(state, "perRiffle", b1)
        ss.preload_metric_phase4_workspace(state, "perRiffle")
        phase4 = ss.get_metric_phase4_display_state(state, "perRiffle")
        # simulate the M5 "finalize" that populates completed_metrics for DEEP
        state.completed_metrics.set(
            {"perRiffle": {"phase4_curve_rows": phase4["curve_rows"]}}
        )
    return state


def test_export_context_has_curve_for_finalized_metric(export_state):
    with reactive.isolate():
        ctx = ss.build_summary_export_context(export_state, metrics=["perRiffle"])
    assert set(ctx) >= {"session_meta", "threshold_rows", "metrics", "discipline_function_mapping"}
    assert len(ctx["threshold_rows"]) >= 1
    assert "metric" in ctx["threshold_rows"].columns
    assert "perRiffle" in ctx["metrics"]
    assert callable(ctx["metrics"]["perRiffle"]["plot_png"])


def test_list_of_metrics_export_from_real_context(export_state):
    with reactive.isolate():
        ctx = ss.build_summary_export_context(export_state, metrics=["perRiffle"])
        mc = export_state.metric_config()
        mapping = export_state.discipline_function_mapping()
    data = _xlsx_to_bytes(
        lambda out: build_oh_list_of_metrics(
            ctx,
            str(TEMPLATES_DIR / "MN-List-of-Metricsv2.0.xlsx"),
            out,
            metric_config=mc,
            mapping=mapping,
        )
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))  # a valid workbook
    assert "Performance Standards" in [s for s in wb.sheetnames]


def test_sqt_workbook_export_from_real_context(export_state):
    with reactive.isolate():
        ctx = ss.build_summary_export_context(export_state, metrics=["perRiffle"])
        mc = export_state.metric_config()
        sc = export_state.strat_config()
        mapping = export_state.discipline_function_mapping()
    data = _xlsx_to_bytes(
        lambda out: build_oh_reference_curves_workbook(
            ctx, out, metric_config=mc, strat_config=sc, mapping=mapping
        )
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.active is not None


def test_science_report_from_real_context(export_state):
    with reactive.isolate():
        ctx = ss.build_summary_export_context(export_state, metrics=["perRiffle"])
        mc = export_state.metric_config()
    ctx["session_meta"]["generated_at"] = "2026-07-05 12:00:00"
    html = build_science_support_html(ctx, mc)
    assert html.startswith("<!DOCTYPE html>")
    assert "Percent Riffle" in html or "perRiffle" in html


def test_deep_collect_curve_rows_from_finalized_curve(export_state):
    # the new integration point: the finalized completed_metrics entry is
    # collected into per-metric curve rows (the full bundle build against a
    # canonical STAF mapping is golden-tested in test_deep_export).
    with reactive.isolate():
        completed = export_state.completed_metrics()
    curve_rows = deep_collect_curve_rows(completed)
    assert "perRiffle" in curve_rows
    assert deep_collect_curve_rows({}) == {}  # the "no finalized curves" gate
