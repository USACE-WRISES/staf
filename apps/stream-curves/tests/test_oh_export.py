"""Tests for streamcurves.oh_export (port of R/14_oh_list_of_metrics.R and
R/15_oh_sqt_workbook.R).

The end-to-end fixtures mirror tests/oh_output_builders_tests.R from the R
repo (mock_threshold_rows / mock_metric_config / mock_context) so golden
workbooks exported by R with the same mocks are directly comparable.
"""

from __future__ import annotations

import logging
import math
from io import BytesIO
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from streamcurves import oh_export as ohx
from streamcurves.paths import TEMPLATES_DIR

LOM_TEMPLATE = TEMPLATES_DIR / "MN-List-of-Metricsv2.0.xlsx"
GOLDEN_DIR = Path(__file__).parent / "golden"


# --------------------------------------------------------------------------- #
# Fixtures (verbatim port of the R test mocks)
# --------------------------------------------------------------------------- #


def r_mock_threshold_rows() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "metric": [
                "perRiffle", "wBHR", "ER", "LWD_frequency", "Per_erodingbank",
                "Avgherbcover", "avgBasalarea", "median_PSR", "avg_PDR", "d50",
                "WDR",
                # non-OH metric to confirm filtering:
                "fakeMetricNotInMap",
            ],
            "stratum": ["StreamType2 = B"] + [None] * 11,
            "higher_is_better": [
                True, False, True, True, False, True, True, False, True, None,
                False, True,
            ],
            "not_functioning_min": [0, 2.0, 1.0, 0, 75, 10, 2, 8, 1.1, np.nan, 20, np.nan],
            "not_functioning_max": [20, np.inf, 1.5, 30, np.inf, 55, 10, 6.5, 1.8, np.nan, np.inf, np.nan],
            "at_risk_min": [20, 1.3, 1.5, 30, 30, 55, 10, 6.5, 1.8, np.nan, 12, np.nan],
            "at_risk_max": [41, 2.0, 2.2, 45, 75, 80, 12, 5, 2.5, np.nan, 20, np.nan],
            "functioning_min": [41, 1.0, 2.2, 45, 5, 80, 12, 5, 2.5, np.nan, 7, np.nan],
            "functioning_max": [60, 1.3, np.inf, 60, 30, 100, 14, 4, 3, np.nan, 12, np.nan],
            "curve_point1_x": [0, 2.0, 1.0, 0, 75, 10, 2, 8, 1.1, 2, 20, 0],
            "curve_point1_y": [0.0] * 12,
            "curve_point2_x": [40, 1.3, 1.5, 30, 30, 55, 10, 6.5, 1.8, 15, 12, 0],
            "curve_point2_y": [0.7] * 12,
            "curve_point3_x": [60, 1.0, 2.5, 60, 5, 80, 14, 5, 2.5, 40, 7, 0],
            "curve_point3_y": [1.0] * 12,
        }
    )


def r_mock_metric_config() -> dict:
    return {
        "perRiffle": {"display_name": "Percent Riffle", "units": "%", "higher_is_better": True},
        "wBHR": {"display_name": "Bank Height Ratio", "units": "ratio", "higher_is_better": False},
        "ER": {"display_name": "Entrenchment Ratio", "units": "ratio", "higher_is_better": True},
        "LWD_frequency": {"display_name": "LWD Frequency", "units": "count/reach", "higher_is_better": True},
        "Per_erodingbank": {"display_name": "Percent Streambank Erosion", "units": "%", "higher_is_better": False},
        "Avgherbcover": {"display_name": "Native Herbaceous Cover", "units": "%", "higher_is_better": True},
        "avgBasalarea": {"display_name": "Woody Stem Basal Area", "units": "sqft/acre", "higher_is_better": True},
        "median_PSR": {"display_name": "Pool Spacing Ratio", "units": "ratio", "higher_is_better": False},
        "avg_PDR": {"display_name": "Pool Depth Ratio", "units": "ratio", "higher_is_better": True},
        "d50": {"display_name": "Median Grain Size (d50)", "units": "mm", "higher_is_better": None},
        "WDR": {"display_name": "Width/Depth Ratio", "units": "ratio", "higher_is_better": False},
    }


def r_mock_context() -> dict:
    return {
        "threshold_rows": r_mock_threshold_rows(),
        "metrics": {},
        "session_meta": {
            "generated_at": "2026-04-16 test",
            "metric_count": 11,
            "complete_metrics": 11,
            "review_metrics": 0,
            "manual_curve_metrics": 0,
        },
    }


def r_mock_strat_config() -> dict:
    return {"StreamType2": {"levels": ["B", "C", "E", "F"]}}


def _fill(cell) -> str:
    return cell.fill.start_color.rgb


def _all_string_values(ws) -> list[str]:
    out = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                out.append(cell.value)
    return out


# --------------------------------------------------------------------------- #
# Pure helpers: index values
# --------------------------------------------------------------------------- #


class TestIndexValues:
    def test_two_points(self):
        assert ohx.oh_sqt_index_values(2) == [0.0, 1.0]
        assert ohx.oh_sqt_index_values(2, higher_is_better=False) == [1.0, 0.0]

    def test_three_points(self):
        assert ohx.oh_sqt_index_values(3) == [0.0, 0.70, 1.00]
        assert ohx.oh_sqt_index_values(3, higher_is_better=False) == [1.00, 0.70, 0.0]

    def test_four_points(self):
        assert ohx.oh_sqt_index_values(4) == [0.0, 0.30, 0.70, 1.00]
        assert ohx.oh_sqt_index_values(4, higher_is_better=False) == [1.00, 0.70, 0.30, 0.0]

    def test_seven_points_linspace(self):
        got = ohx.oh_sqt_index_values(7)
        assert got == pytest.approx(list(np.linspace(0, 1, 7)))
        got_rev = ohx.oh_sqt_index_values(7, higher_is_better=False)
        assert got_rev == pytest.approx(list(np.linspace(0, 1, 7))[::-1])

    def test_degenerate_counts(self):
        assert ohx.oh_sqt_index_values(0) == []
        assert ohx.oh_sqt_index_values(-2) == []
        assert ohx.oh_sqt_index_values(1) == [0.0]  # seq(0, 1, length.out = 1)

    def test_na_higher_is_better_reverses(self):
        # R: isTRUE(NA) is FALSE -> reversed
        assert ohx.oh_sqt_index_values(3, higher_is_better=None) == [1.00, 0.70, 0.0]


# --------------------------------------------------------------------------- #
# Pure helpers: linear segments
# --------------------------------------------------------------------------- #


class TestLinearSegments:
    def test_known_three_point_curve(self):
        segs = ohx.oh_sqt_linear_segments([0, 40, 60], [0, 0.7, 1])
        assert segs["slopes"] == pytest.approx([0.0175, 0.015])
        assert segs["intercepts"] == pytest.approx([0.0, 0.1])
        assert segs["seg_labels"] == ["0 → 0.7", "0.7 → 1"]

    def test_lower_is_better_curve(self):
        # Field values sorted by index score ascending with reversed iv.
        segs = ohx.oh_sqt_linear_segments([75, 30, 5], [1, 0.7, 0])
        assert segs["slopes"] == pytest.approx([(0.7 - 1) / (30 - 75), (0 - 0.7) / (5 - 30)])
        assert segs["intercepts"] == pytest.approx([1 - (0.3 / 45) * 75, 0.7 - 0.028 * 30])
        assert segs["seg_labels"] == ["1 → 0.7", "0.7 → 0"]

    def test_zero_width_segment_is_nan(self):
        segs = ohx.oh_sqt_linear_segments([10, 10], [0, 1])
        assert len(segs["slopes"]) == 1
        assert math.isnan(segs["slopes"][0])
        assert math.isnan(segs["intercepts"][0])
        assert segs["seg_labels"] == ["0 → 1"]

    def test_single_point_empty(self):
        segs = ohx.oh_sqt_linear_segments([5], [0])
        assert segs == {"seg_labels": [], "slopes": [], "intercepts": []}


# --------------------------------------------------------------------------- #
# Pure helpers: band labels
# --------------------------------------------------------------------------- #


class TestBandLabels:
    def test_segment_band_labels_three_point(self):
        # Segment 0.7 -> 1 is "F" only: the FAR guard is y_lo < 0.70 (strict),
        # so 0.7 does not re-enter the FAR band (R/15 oh_sqt_segment_band_labels).
        assert ohx.oh_sqt_segment_band_labels([0, 0.7, 1]) == ["NF/FAR", "F"]

    def test_segment_band_labels_four_point(self):
        assert ohx.oh_sqt_segment_band_labels([0, 0.3, 0.7, 1]) == ["NF", "FAR", "F"]

    def test_segment_band_labels_reversed(self):
        assert ohx.oh_sqt_segment_band_labels([1, 0.7, 0]) == ["F", "NF/FAR"]

    def test_segment_band_labels_flat_segments(self):
        assert ohx.oh_sqt_segment_band_labels([0.5, 0.5]) == ["FAR"]
        assert ohx.oh_sqt_segment_band_labels([0, 0]) == ["—"]

    def test_segment_band_labels_degenerate(self):
        assert ohx.oh_sqt_segment_band_labels([0.5]) == []
        assert ohx.oh_sqt_segment_band_labels([]) == []

    def test_oh_sqt_band(self):
        assert ohx.oh_sqt_band(0.0) == "NF"
        assert ohx.oh_sqt_band(0.29) == "NF"
        assert ohx.oh_sqt_band(0.30) == "FAR"
        assert ohx.oh_sqt_band(0.69) == "FAR"
        assert ohx.oh_sqt_band(0.70) == "F"
        assert ohx.oh_sqt_band(1.0) == "F"
        assert ohx.oh_sqt_band(float("nan")) is None


# --------------------------------------------------------------------------- #
# Pure helpers: threshold cell formatting
# --------------------------------------------------------------------------- #


class TestFormatThresholdCell:
    def test_signif_three_non_scientific(self):
        assert ohx.format_threshold_cell(41) == "41"
        assert ohx.format_threshold_cell(0.001234) == "0.00123"
        assert ohx.format_threshold_cell(1234567) == "1230000"
        assert ohx.format_threshold_cell(999.9) == "1000"
        assert ohx.format_threshold_cell(2 / 3) == "0.667"
        assert ohx.format_threshold_cell(1.3) == "1.3"
        assert ohx.format_threshold_cell(0) == "0"
        assert ohx.format_threshold_cell(-0.001234) == "-0.00123"

    def test_non_finite_is_dash(self):
        assert ohx.format_threshold_cell(None) == "-"
        assert ohx.format_threshold_cell(float("nan")) == "-"
        assert ohx.format_threshold_cell(np.nan) == "-"
        assert ohx.format_threshold_cell(float("inf")) == "-"
        assert ohx.format_threshold_cell(-np.inf) == "-"
        # R: is.finite(<character>) is FALSE -> "-"
        assert ohx.format_threshold_cell("41") == "-"

    def test_boundary_prefixes(self):
        assert ohx.format_threshold_cell(41, "min", True, boundary=True) == "≥ 41"
        assert ohx.format_threshold_cell(41, "max", True, boundary=True) == "≤ 41"
        # NOTE(parity): R computes the same prefix for both directions.
        assert ohx.format_threshold_cell(41, "min", False, boundary=True) == "≥ 41"
        assert ohx.format_threshold_cell(41, "max", False, boundary=True) == "≤ 41"

    def test_invalid_side_raises(self):
        with pytest.raises(ValueError):
            ohx.format_threshold_cell(1, side="mid")


# --------------------------------------------------------------------------- #
# Row-values / stratum parsing / points extraction
# --------------------------------------------------------------------------- #


class TestRowValuesAndPoints:
    def test_row_values_stratum_split(self):
        row = r_mock_threshold_rows().iloc[[0]]  # perRiffle, StreamType2 = B
        values = ohx.oh_list_of_metrics_row_values(
            "perRiffle", row, metric_config=r_mock_metric_config()
        )
        assert values["metric_label"] == "Percent Riffle (%)"
        assert values["strat_type"] == "StreamType2"
        assert values["strat_desc"] == "B"
        assert values["nf_min"] == "0"
        assert values["nf_max"] == "20"
        assert values["f_max"] == "60"
        assert values["applicability"] == ""
        assert values["notes"].startswith("Percent of reach length")

    def test_row_values_colon_stratum_and_none(self):
        base = r_mock_threshold_rows().iloc[[0]].copy()
        base["stratum"] = "Region : North : East"
        values = ohx.oh_list_of_metrics_row_values("perRiffle", base)
        assert values["strat_type"] == "Region"
        assert values["strat_desc"] == "North East"

        base["stratum"] = "none"
        values = ohx.oh_list_of_metrics_row_values("perRiffle", base)
        assert values["strat_type"] == ""
        assert values["strat_desc"] == ""

        base["stratum"] = "plainlabel"
        values = ohx.oh_list_of_metrics_row_values("perRiffle", base)
        assert values["strat_type"] == ""
        assert values["strat_desc"] == "plainlabel"

    def test_row_values_inf_thresholds_dash(self):
        rows = r_mock_threshold_rows()
        row = rows[rows["metric"] == "wBHR"]
        values = ohx.oh_list_of_metrics_row_values("wBHR", row, r_mock_metric_config())
        assert values["nf_min"] == "2"
        assert values["nf_max"] == "-"  # Inf
        assert values["far_min"] == "1.3"

    def test_points_sorted_by_index_score(self):
        rows = r_mock_threshold_rows()
        assert ohx.oh_sqt_ohio_points_for_metric(rows, "perRiffle") == [0.0, 40.0, 60.0]
        # lower-is-better metric: x's still sorted by index score ascending
        assert ohx.oh_sqt_ohio_points_for_metric(rows, "Per_erodingbank") == [75.0, 30.0, 5.0]
        assert ohx.oh_sqt_ohio_points_for_metric(rows, "notARealMetric") == []
        assert ohx.oh_sqt_ohio_points_for_metric(None, "perRiffle") == []

    def test_points_from_nested_curve_points(self):
        df = pd.DataFrame({"metric": ["m1"], "higher_is_better": [True]})
        cps = np.empty(1, dtype=object)
        cps[0] = [
            {"metric_value": 30.0, "index_score": 0.7},
            {"metric_value": 10.0, "index_score": 0.0},
            {"metric_value": 20.0, "index_score": 0.3},
            {"metric_value": 40.0, "index_score": 1.0},
        ]
        df["curve_points"] = cps
        assert ohx.oh_sqt_ohio_points_for_metric(df, "m1") == [10.0, 20.0, 30.0, 40.0]

    def test_threshold_rows_for_metric(self):
        rows = r_mock_threshold_rows()
        got = ohx.oh_threshold_rows_for_metric(rows, "perRiffle")
        assert len(got) == 1
        assert got["stratum"].iloc[0] == "StreamType2 = B"
        assert len(ohx.oh_threshold_rows_for_metric(None, "perRiffle")) == 0
        assert len(ohx.oh_threshold_rows_for_metric(pd.DataFrame(), "perRiffle")) == 0
        no_metric_col = pd.DataFrame({"x": [1]})
        assert len(ohx.oh_threshold_rows_for_metric(no_metric_col, "perRiffle")) == 0


# --------------------------------------------------------------------------- #
# End-to-end: List of Metrics (real bundled template)
# --------------------------------------------------------------------------- #


class TestBuildListOfMetrics:
    @pytest.fixture(scope="class")
    def built(self, tmp_path_factory):
        out = tmp_path_factory.mktemp("lom") / "list_of_metrics.xlsx"
        ohx.build_oh_list_of_metrics(
            context=r_mock_context(),
            template_path=LOM_TEMPLATE,
            output_path=out,
            metric_config=r_mock_metric_config(),
        )
        return openpyxl.load_workbook(out)

    def test_sheets_preserved(self, built):
        assert built.sheetnames == ["Performance Standards", "References"]

    def test_hydraulics_rows_first(self, built):
        ws = built["Performance Standards"]
        # data order within category: wBHR, ER, WDR (rows 3-5)
        assert ws.cell(3, 1).value == "Hydraulics"
        assert ws.cell(3, 2).value == "Floodplain Connectivity"
        assert ws.cell(3, 3).value == "Bank Height Ratio (ratio)"
        assert ws.cell(3, 6).value == "2"
        assert ws.cell(3, 7).value == "-"  # Inf -> "-"
        assert ws.cell(3, 8).value == "1.3"
        assert ws.cell(3, 9).value == "2"
        assert ws.cell(3, 10).value == "1"
        assert ws.cell(3, 11).value == "1.3"
        assert ws.cell(4, 3).value == "Entrenchment Ratio (ratio)"
        assert ws.cell(5, 3).value == "Width/Depth Ratio (ratio)"
        assert ws.cell(5, 2).value == "Width/Depth Ratio"

    def test_geomorphology_rows_and_stratum_split(self, built):
        ws = built["Performance Standards"]
        # geomorph starts at row 6 with perRiffle (data order)
        assert ws.cell(6, 1).value == "Geomorphology"
        assert ws.cell(6, 2).value == "Bedform Diversity"
        assert ws.cell(6, 3).value == "Percent Riffle (%)"
        assert ws.cell(6, 4).value == "StreamType2"
        assert ws.cell(6, 5).value == "B"
        assert ws.cell(6, 13).value.startswith("Percent of reach length")
        # d50 has NA thresholds -> all "-"
        assert ws.cell(13, 3).value == "Median Grain Size (d50) (mm)"
        assert [ws.cell(13, c).value for c in range(6, 12)] == ["-"] * 6

    def test_category_on_every_row(self, built):
        ws = built["Performance Standards"]
        for r in range(6, 14):
            assert ws.cell(r, 1).value == "Geomorphology"

    def test_unmapped_metric_absent(self, built):
        ws = built["Performance Standards"]
        body = [
            str(ws.cell(r, c).value)
            for r in range(3, 55)
            for c in range(1, 14)
            if ws.cell(r, c).value is not None
        ]
        assert not any("fakeMetric" in v for v in body)

    def test_rows_after_last_metric_cleared(self, built):
        ws = built["Performance Standards"]
        for c in range(1, 14):
            assert ws.cell(20, c).value in (None, "")

    def test_notes_footer_preserved(self, built):
        ws = built["Performance Standards"]
        assert ws.cell(55, 1).value == "Notes:"
        assert "reference standards" in ws.cell(55, 2).value

    def test_references_sheet(self, built):
        ws = built["References"]
        assert ws.cell(1, 1).value == "Functional Category"  # header untouched
        # YAML coverage: Hydrology (CFPs) first
        assert ws.cell(2, 1).value == "Hydrology"
        assert ws.cell(2, 2).value == "Reach Runoff"
        assert ws.cell(2, 3).value == "CFPs"  # no display_name in metric_config
        assert ws.cell(2, 4).value == "OSAM reference-stream dataset"
        # Hydraulics rows 3-6 in YAML order: WDR, wBHR, ER, wER
        assert ws.cell(3, 1).value == "Hydraulics"
        assert ws.cell(3, 3).value == "Width/Depth Ratio"
        assert ws.cell(4, 1).value in (None, "")  # category only on first row
        assert ws.cell(4, 3).value == "Bank Height Ratio"
        # BEHI_NBS (14th Geomorphology metric -> row 20) joins two sources
        assert ws.cell(20, 4).value == (
            "Rosgen, D.L. (2001) BEHI/NBS methodology; OSAM reference-stream dataset"
        )

    def test_template_missing_raises(self, tmp_path):
        with pytest.raises(ValueError, match="template not found"):
            ohx.build_oh_list_of_metrics(
                r_mock_context(), tmp_path / "nope.xlsx", tmp_path / "out.xlsx"
            )


class TestListOfMetricsMappingAndTruncation:
    def test_mapping_override_orders_by_function_label(self, tmp_path):
        mapping = pd.DataFrame(
            {
                "metric_key": ["WDR", "wBHR", "ER"],
                "discipline": ["Hydraulics"] * 3,
                "function_label": [
                    "Custom Hydraulics Fn",
                    "Floodplain Connectivity",
                    "Floodplain Connectivity",
                ],
                "sort_order": [1, 2, 3],
            }
        )
        out = tmp_path / "lom_mapped.xlsx"
        ohx.build_oh_list_of_metrics(
            context=r_mock_context(),
            template_path=LOM_TEMPLATE,
            output_path=out,
            metric_config=r_mock_metric_config(),
            mapping=mapping,
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["Performance Standards"]
        # ordered by (function_label, sort_order): WDR first
        assert ws.cell(3, 2).value == "Custom Hydraulics Fn"
        assert ws.cell(3, 3).value == "Width/Depth Ratio (ratio)"
        assert ws.cell(4, 3).value == "Bank Height Ratio (ratio)"
        assert ws.cell(5, 3).value == "Entrenchment Ratio (ratio)"
        # mapping restricts both sheets to its own metric keys
        for c in range(1, 14):
            assert ws.cell(6, c).value in (None, "")
        refs = wb["References"]
        # three mapped metrics -> rows 2..4 (WDR, wBHR, ER)
        assert refs.cell(2, 1).value == "Hydraulics"
        assert refs.cell(2, 2).value == "Custom Hydraulics Fn"
        assert refs.cell(2, 3).value == "Width/Depth Ratio"
        assert refs.cell(3, 1).value in (None, "")  # category only on first row
        assert refs.cell(3, 3).value == "Bank Height Ratio"
        assert refs.cell(4, 3).value == "Entrenchment Ratio"
        for c in range(1, 5):
            assert refs.cell(5, c).value in (None, "")

    def test_truncation_past_row_54_warns(self, tmp_path, caplog):
        n = 60
        rows = pd.DataFrame(
            {
                "metric": ["perRiffle"] * n,
                "stratum": [f"StreamType2 = S{i}" for i in range(n)],
                "higher_is_better": [True] * n,
                "not_functioning_min": [0.0] * n,
                "not_functioning_max": [20.0] * n,
                "at_risk_min": [20.0] * n,
                "at_risk_max": [41.0] * n,
                "functioning_min": [41.0] * n,
                "functioning_max": [60.0] * n,
            }
        )
        out = tmp_path / "lom_trunc.xlsx"
        with caplog.at_level(logging.WARNING, logger="streamcurves"):
            ohx.build_oh_list_of_metrics(
                context={"threshold_rows": rows},
                template_path=LOM_TEMPLATE,
                output_path=out,
                metric_config=r_mock_metric_config(),
            )
        assert "truncating at row 54" in caplog.text

        wb = openpyxl.load_workbook(out)
        ws = wb["Performance Standards"]
        # rows 3..54 hold strata S0..S51; nothing spills past row 54
        assert ws.cell(3, 5).value == "S0"
        assert ws.cell(54, 5).value == "S51"
        assert ws.cell(55, 1).value == "Notes:"


# --------------------------------------------------------------------------- #
# End-to-end: Reference Curves workbook (R-mock parity, YAML fallback)
# --------------------------------------------------------------------------- #


class TestBuildReferenceCurvesWorkbook:
    @pytest.fixture(scope="class")
    def built(self, tmp_path_factory):
        out = tmp_path_factory.mktemp("sqt") / "reference_curves.xlsx"
        ohx.build_oh_reference_curves_workbook(
            context=r_mock_context(),
            output_path=out,
            metric_config=r_mock_metric_config(),
            strat_config=r_mock_strat_config(),
            bundle_dir=None,
        )
        return openpyxl.load_workbook(out)

    def test_sheet_names(self, built):
        assert built.sheetnames == ["Reference Standards", "Pull Down Notes"]

    def test_title_and_subtitle(self, built):
        ws = built["Reference Standards"]
        assert ws["B1"].value == "Stream Quantification Tool — Reference Curves"
        assert _fill(ws["B1"]) == "FF1F4E79"
        assert ws["B1"].font.bold and ws["B1"].font.size == 16
        assert "B1:AO1" in {str(r) for r in ws.merged_cells.ranges}
        assert ws.row_dimensions[1].height == 28
        assert ws["B2"].value == (
            "Session: 2026-04-16 test  |  Complete metrics: 11"
            "  |  Performance bands: NF < 0.30 ≤ FAR < 0.70 ≤ F"
        )

    def test_column_widths(self, built):
        ws = built["Reference Standards"]
        assert ws.column_dimensions["A"].width == 2
        assert ws.column_dimensions["B"].width == 16
        assert ws.column_dimensions["C"].width == 9
        assert ws.column_dimensions["I"].width == 3
        assert ws.column_dimensions["J"].width == 16  # second band label col

    def test_discipline_headers_row5(self, built):
        ws = built["Reference Standards"]
        expected = {
            "B5": ("HYDROLOGY", "FFDCE6F1"),
            "J5": ("HYDRAULICS", "FFB4C7E7"),
            "R5": ("GEOMORPHOLOGY", "FFFBE5D6"),
            "Z5": ("PHYSICOCHEMISTRY", "FFFFF2CC"),
            "AH5": ("BIOLOGY", "FFE2EFDA"),
        }
        for ref, (text, fill) in expected.items():
            assert ws[ref].value == text
            assert _fill(ws[ref]) == fill
        merges = {str(r) for r in ws.merged_cells.ranges}
        assert {"B5:I5", "J5:Q5", "R5:Y5", "Z5:AG5", "AH5:AO5"} <= merges
        assert ws.row_dimensions[5].height == 24

    def test_hydraulics_first_block_lower_is_better(self, built):
        ws = built["Reference Standards"]
        # wBHR: no stratum, lower-is-better
        assert ws["J8"].value == "Bank Height Ratio (ratio)"
        assert _fill(ws["J8"]) == "FFD9E1F2"
        assert "J8:Q8" in {str(r) for r in ws.merged_cells.ranges}
        assert ws.row_dimensions[8].height == 22
        assert ws["J9"].value == "Floodplain Connectivity"
        assert ws["J10"].value == "Field Value"
        # x's sorted by index score ascending
        assert [ws["K10"].value, ws["L10"].value, ws["M10"].value] == [2.0, 1.3, 1.0]
        assert ws["J11"].value == "Index Value"
        # reversed index values for lower-is-better (R quirk preserved)
        assert [ws["K11"].value, ws["L11"].value, ws["M11"].value] == [1.0, 0.7, 0.0]
        assert _fill(ws["K11"]) == "FFC6E0B4"  # 1.0 -> green
        assert _fill(ws["L11"]) == "FFC6E0B4"  # 0.7 -> green
        assert _fill(ws["M11"]) == "FFF8CBAD"  # 0.0 -> red
        assert ws["K11"].number_format == "0.##"
        # segment band labels
        assert ws["K13"].value == "F"
        assert ws["L13"].value == "NF/FAR"
        # coefficients
        assert ws["J14"].value == "Coefficients — Y = a * X + b"
        assert "J14:Q14" in {str(r) for r in ws.merged_cells.ranges}
        assert ws["J15"].value == "a"
        assert ws["J16"].value == "b"
        assert ws["K15"].value == pytest.approx((0.7 - 1) / (1.3 - 2.0))
        assert ws["K16"].value == pytest.approx(1 - ((0.7 - 1) / (1.3 - 2.0)) * 2.0)
        assert ws["L15"].value == pytest.approx((0 - 0.7) / (1.0 - 1.3))
        assert ws["K15"].number_format == "0.####"

    def test_block_stacking_within_band(self, built):
        ws = built["Reference Standards"]
        # Hydraulics band: wBHR (8), ER (32), WDR (56) — 22-row block + 2 gap
        assert ws["J32"].value == "Entrenchment Ratio (ratio)"
        assert ws["J56"].value == "Width/Depth Ratio (ratio)"

    def test_geomorphology_block_with_stratum(self, built):
        ws = built["Reference Standards"]
        assert ws["R8"].value == "Percent Riffle (%) — StreamType2 = B"
        assert [ws["S10"].value, ws["T10"].value, ws["U10"].value] == [0.0, 40.0, 60.0]
        assert [ws["S11"].value, ws["T11"].value, ws["U11"].value] == [0.0, 0.7, 1.0]
        assert _fill(ws["S11"]) == "FFF8CBAD"
        assert _fill(ws["T11"]) == "FFC6E0B4"
        assert ws["S13"].value == "NF/FAR"
        assert ws["T13"].value == "F"  # FAR guard is y_lo < 0.70 (strict)
        assert ws["S15"].value == pytest.approx(0.0175)
        assert ws["T15"].value == pytest.approx(0.015)
        assert ws["S16"].value == pytest.approx(0.0)
        assert ws["T16"].value == pytest.approx(0.1)

    def test_na_higher_is_better_reverses_iv(self, built):
        ws = built["Reference Standards"]
        # d50 is the 8th Geomorphology block: row 8 + 7*24 = 176
        assert ws["R176"].value == "Median Grain Size (d50) (mm)"
        assert [ws["S178"].value, ws["T178"].value, ws["U178"].value] == [2.0, 15.0, 40.0]
        assert [ws["S179"].value, ws["T179"].value, ws["U179"].value] == [1.0, 0.7, 0.0]

    def test_unmapped_metric_absent(self, built):
        values = _all_string_values(built["Reference Standards"])
        assert not any("fakeMetric" in v for v in values)

    def test_every_data_metric_has_a_block(self, built):
        values = _all_string_values(built["Reference Standards"])
        config = r_mock_metric_config()
        for metric_key in config:
            display = config[metric_key]["display_name"]
            assert any(display in v for v in values), metric_key
        n_blocks = sum(v == "Field Value" for v in values)
        assert n_blocks == 11
        assert sum(v == "Index Value" for v in values) == 11
        assert sum(v.startswith("Coefficients") for v in values) == 11

    def test_pull_down_notes(self, built):
        ws = built["Pull Down Notes"]
        assert ws["A1"].value == "Stream Type:"
        assert [ws.cell(r, 2).value for r in range(2, 6)] == ["B", "C", "E", "F"]
        assert ws.column_dimensions["A"].width == 14
        assert ws.column_dimensions["B"].width == 18


# --------------------------------------------------------------------------- #
# End-to-end: Reference Curves with mapping, nested points, plots
# --------------------------------------------------------------------------- #


def _png_bytes() -> bytes:
    from PIL import Image as PILImage

    buf = BytesIO()
    PILImage.new("RGB", (20, 13), (200, 30, 30)).save(buf, format="PNG")
    return buf.getvalue()


class TestReferenceCurvesMappingAndPlots:
    @pytest.fixture(scope="class")
    def built(self, tmp_path_factory):
        rows = pd.DataFrame(
            {
                "metric": ["perRiffle", "perRiffle", "Per_erodingbank", "d4metric"],
                "stratum": ["StreamType2 = B", "StreamType2 = C", None, None],
                "higher_is_better": [True, True, False, True],
                "curve_point1_x": [0.0, 5.0, 75.0, np.nan],
                "curve_point1_y": [0.0, 0.0, 0.0, np.nan],
                "curve_point2_x": [40.0, 45.0, 30.0, np.nan],
                "curve_point2_y": [0.7, 0.7, 0.7, np.nan],
                "curve_point3_x": [60.0, 65.0, 5.0, np.nan],
                "curve_point3_y": [1.0, 1.0, 1.0, np.nan],
            }
        )
        cps = np.empty(4, dtype=object)
        cps[3] = [
            {"metric_value": 10.0, "index_score": 0.0},
            {"metric_value": 20.0, "index_score": 0.3},
            {"metric_value": 30.0, "index_score": 0.7},
            {"metric_value": 40.0, "index_score": 1.0},
        ]
        rows["curve_points"] = cps
        mapping = pd.DataFrame(
            {
                "metric_key": ["perRiffle", "Per_erodingbank", "d4metric"],
                "discipline": ["Hydrology", "Hydraulics", "Geomorphology"],
                "function_label": ["Fn Riffle", "Fn Erosion", "Fn Substrate"],
                "sort_order": [1, 2, 3],
            }
        )
        metric_config = {
            "perRiffle": {"display_name": "Percent Riffle", "units": "%"},
            "Per_erodingbank": {"display_name": "Percent Streambank Erosion", "units": "%"},
            "d4metric": {"display_name": "D4 Metric", "units": "mm"},
        }
        context = {
            "threshold_rows": rows,
            "metrics": {"perRiffle": {"plot_png": _png_bytes()}},
            "session_meta": {"generated_at": "2026-07-02 test", "complete_metrics": 3},
        }
        out = tmp_path_factory.mktemp("sqt_mapped") / "reference_curves_mapped.xlsx"
        ohx.build_oh_reference_curves_workbook(
            context=context,
            output_path=out,
            metric_config=metric_config,
            strat_config=None,  # falls back to B/C/E/F
            bundle_dir=None,
            mapping=mapping,
        )
        return openpyxl.load_workbook(out)

    def test_mapping_places_disciplines(self, built):
        ws = built["Reference Standards"]
        assert ws["B8"].value == "Percent Riffle (%) — StreamType2 = B"
        assert ws["B9"].value == "Fn Riffle"
        assert ws["J8"].value == "Percent Streambank Erosion (%)"
        assert ws["J9"].value == "Fn Erosion"
        assert ws["R8"].value == "D4 Metric (mm)"

    def test_one_block_per_stratum(self, built):
        ws = built["Reference Standards"]
        assert ws["B32"].value == "Percent Riffle (%) — StreamType2 = C"
        assert [ws["C34"].value, ws["D34"].value, ws["E34"].value] == [5.0, 45.0, 65.0]

    def test_nested_curve_points_four_point_block(self, built):
        ws = built["Reference Standards"]
        assert [ws.cell(10, c).value for c in range(19, 23)] == [10.0, 20.0, 30.0, 40.0]
        assert [ws.cell(11, c).value for c in range(19, 23)] == [0.0, 0.3, 0.7, 1.0]
        assert _fill(ws.cell(11, 19)) == "FFF8CBAD"  # 0.0 red
        assert _fill(ws.cell(11, 20)) == "FFFFE699"  # 0.3 yellow
        assert _fill(ws.cell(11, 21)) == "FFC6E0B4"  # 0.7 green
        assert _fill(ws.cell(11, 22)) == "FFC6E0B4"  # 1.0 green
        assert [ws.cell(13, c).value for c in range(19, 22)] == ["NF", "FAR", "F"]

    def test_plot_images_embedded_per_stratum_block(self, built):
        ws = built["Reference Standards"]
        assert len(ws._images) == 2  # one per perRiffle stratum block
        anchors = {(img.anchor._from.row, img.anchor._from.col) for img in ws._images}
        assert anchors == {(17, 1), (41, 1)}  # 0-based B18 and B42

    def test_pull_down_notes_fallback_levels(self, built):
        ws = built["Pull Down Notes"]
        assert [ws.cell(r, 2).value for r in range(2, 6)] == ["B", "C", "E", "F"]

    def test_unmapped_disciplines_have_no_bands(self, built):
        # With a mapping present, resolved_category_order() keeps only the
        # ACTIVE disciplines (R/13), so Physicochemistry/Biology get no bands.
        # (With mapping=None all five headers appear — covered by
        # TestBuildReferenceCurvesWorkbook.test_discipline_headers_row5.)
        ws = built["Reference Standards"]
        assert ws["B5"].value == "HYDROLOGY"
        assert ws["J5"].value == "HYDRAULICS"
        assert ws["R5"].value == "GEOMORPHOLOGY"
        assert ws["Z5"].value is None
        assert ws["AH5"].value is None


# --------------------------------------------------------------------------- #
# Golden-fixture comparisons (skip when fixtures are absent)
# --------------------------------------------------------------------------- #


def _norm_cell(v):
    if v is None or v == "":
        return None
    return v


def _assert_workbook_values_match(golden_path: Path, ours_path: Path, sheet: str,
                                  max_row: int, max_col: int) -> None:
    golden = openpyxl.load_workbook(golden_path)[sheet]
    ours = openpyxl.load_workbook(ours_path)[sheet]
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            g = _norm_cell(golden.cell(r, c).value)
            o = _norm_cell(ours.cell(r, c).value)
            if isinstance(g, float) and isinstance(o, (int, float)):
                assert o == pytest.approx(g), f"cell ({r},{c})"
            else:
                assert o == g, f"cell ({r},{c})"


@pytest.mark.parametrize(
    "golden_name, builder, sheet, max_row, max_col",
    [
        ("oh_list_of_metrics.xlsx", "lom", "Performance Standards", 54, 13),
        ("oh_reference_curves.xlsx", "sqt", "Reference Standards", 250, 42),
    ],
)
def test_golden_workbooks(tmp_path, golden_name, builder, sheet, max_row, max_col):
    """Compare against R-built workbooks generated with the same mocks
    (scripts/export_golden.R). Skipped until the fixtures land."""
    golden_path = GOLDEN_DIR / golden_name
    if not golden_path.exists():
        pytest.skip(f"golden fixture missing: {golden_path}")
    out = tmp_path / golden_name
    if builder == "lom":
        ohx.build_oh_list_of_metrics(
            r_mock_context(), LOM_TEMPLATE, out, metric_config=r_mock_metric_config()
        )
        start_row = 3
    else:
        ohx.build_oh_reference_curves_workbook(
            r_mock_context(), out,
            metric_config=r_mock_metric_config(),
            strat_config=r_mock_strat_config(),
            bundle_dir=None,
        )
        start_row = 1
    _assert_workbook_values_match(golden_path, out, sheet, max_row, max_col)
    assert start_row <= max_row
