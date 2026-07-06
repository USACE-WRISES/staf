"""Tests for streamcurves.workbook (port of R/00_input_workbook.R + R/01_load_data.R)."""

from __future__ import annotations

import json
import logging
import math
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook

from streamcurves import workbook as wb

FIXTURE = Path(__file__).parent / "fixtures" / "OSAM_summarydata.xlsx"
GOLDEN = Path(__file__).parent / "golden" / "01_bundle_meta.json"

BUNDLE_KEYS = {
    "raw_data",
    "metric_config",
    "strat_config",
    "predictor_config",
    "factor_recode_config",
    "site_mask_config",
    "discipline_function_mapping",
    "mapping_covers_all_metrics",
    "metadata",
}


# --------------------------------------------------------------------------- #
# Synthetic workbook builder
# --------------------------------------------------------------------------- #


def _base_sheets() -> dict[str, list[list]]:
    """Rows (header first) for a small valid workbook."""
    return {
        "data": [
            ["SiteName", "Ecoregion", "Width", "Wood"],
            ["s1", "ECBP", 3.5, 2],
            ["s2", "ECBP", 7.1, 0],
            [None, "HELP", 2.0, 5],
            ["s4", "HELP", 9.9, 1],
            ["s5", "IP", 4.2, 3],
            ["s6", "IP", 5.0, None],
        ],
        "metrics": [
            [
                "metric_key", "display_name", "column_name", "units", "metric_family",
                "higher_is_better", "monotonic_linear", "preferred_transform",
                "min_sample_size", "best_subsets_allowed", "count_model",
                "stratification_mode", "include_in_summary", "missing_data_rule", "notes",
            ],
            # blank min_sample_size -> default 10; numeric/text/bool flag cells
            ["m_width", "Bankfull Width", "Width", "m", "continuous", "yes", True,
             "log", None, "true", "no", "subset", "y", None, None],
            ["m_wood", None, "Wood", None, None, 1, True, None, 5, True, "1", None,
             "0", "warn", "wood note"],
        ],
        "metric_predictors": [
            ["metric_key", "predictor_key", "sort_order"],
            ["m_width", "p_da", 2],
            ["m_width", "p_slope", 1],
        ],
        "metric_stratifications": [
            ["metric_key", "strat_key", "sort_order"],
            ["m_width", "s_eco", 1],
            ["m_wood", "s_eco", 1],
        ],
        "stratifications": [
            [
                "strat_key", "display_name", "strat_type", "source_column",
                "source_data_type", "primary_strat_key", "secondary_strat_key",
                "derived_column_name", "levels", "pairwise_comparisons",
                "min_group_size", "notes",
            ],
            ["s_eco", "Ecoregion", "raw_single", "Ecoregion", "categorical",
             None, None, None, None, None, 3, None],
            ["s_size", "Size class", "custom_group", "Width", "continuous",
             None, None, None, None, None, None, None],
            ["s_pair", "Eco x Size", "paired", None, None, "s_eco", "s_size",
             None, None, None, None, None],
            ["s_recode", "Eco group", "raw_single", "EcoGroup", "categorical",
             None, None, None, None, None, None, None],
        ],
        "strat_groups": [
            ["strat_key", "group_label", "sort_order", "source_values", "rule_expression"],
            ["s_size", "Big", 2, None, "> 5"],
            ["s_size", "Small", 1, None, "<= 5"],
        ],
        "predictors": [
            [
                "predictor_key", "display_name", "column_name", "type", "derived",
                "derivation_method", "source_columns", "constant", "expected_min",
                "expected_max", "missing_data_rule", "notes",
            ],
            ["p_da", "Drainage area", "DA_mi2", "continuous", "false", None, None,
             None, 0, 100, None, None],
            ["p_slope", None, "Slope", "continuous", True, "ratio", "A|B", 2.59,
             None, None, "warn", None],
        ],
        "factor_recodes": [
            ["recode_key", "source_column", "target_column", "target_level",
             "source_values", "notes"],
            ["rec_eco", "Ecoregion", "EcoGroup", "Agg", "ECBP|HELP", None],
            ["rec_eco", "Ecoregion", "EcoGroup", "Other", "IP", None],
        ],
    }


def _write_xlsx(path: Path, sheets: dict[str, list[list]]) -> Path:
    """Write raw rows straight through openpyxl (independent of the module)."""
    book = Workbook()
    book.remove(book.active)
    for name, rows in sheets.items():
        ws = book.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    book.save(path)
    return path


def _workbook_path(tmp_path: Path, sheets: dict[str, list[list]] | None = None) -> Path:
    return _write_xlsx(tmp_path / "input.xlsx", sheets if sheets is not None else _base_sheets())


def _mapping_tables(rows: list[dict]) -> dict:
    cols = ["discipline", "function_label", "metric_key", "metric_display_name", "notes"]
    return {
        "function_mappings": pd.DataFrame(
            {c: pd.Series([r.get(c) for r in rows], dtype=object) for c in cols}
        )
    }


# --------------------------------------------------------------------------- #
# Cell/vector helper units
# --------------------------------------------------------------------------- #


def test_coerce_flag_variants():
    for truthy in ["true", "T", "1", "yes", "Y", " TRUE ", 1, 1.0, True, np.float64(1.0)]:
        assert wb.coerce_flag(truthy) is True, truthy
    for falsy in ["false", "F", "0", "no", "N", 0, 0.0, False, np.float64(0.0)]:
        assert wb.coerce_flag(falsy, default=True) is False, falsy
    # NA / blank -> default
    assert wb.coerce_flag(None, default=True) is True
    assert wb.coerce_flag(float("nan"), default=False) is False
    assert wb.coerce_flag("", default=True) is True
    assert wb.coerce_flag("   ", default=False) is False


def test_coerce_flag_error():
    with pytest.raises(ValueError, match="Could not parse logical value 'maybe'."):
        wb.coerce_flag("maybe")
    with pytest.raises(ValueError, match="Could not parse logical value '2'."):
        wb.coerce_flag(2.0)


def test_coerce_optional_numeric():
    assert math.isnan(wb.coerce_optional_numeric(None))
    assert math.isnan(wb.coerce_optional_numeric(float("nan")))
    assert wb.coerce_optional_numeric("2.5") == 2.5
    assert wb.coerce_optional_numeric(np.int64(7)) == 7.0
    assert wb.coerce_optional_numeric(True) == 1.0
    with pytest.raises(ValueError, match="Could not parse numeric value 'abc'."):
        wb.coerce_optional_numeric("abc")


def test_parse_pipe_values():
    assert wb.parse_pipe_values("a|b|c") == ["a", "b", "c"]
    assert wb.parse_pipe_values(" a | b ") == ["a", "b"]
    assert wb.parse_pipe_values("a||b") == ["a", "b"]
    assert wb.parse_pipe_values(None) == []
    assert wb.parse_pipe_values(float("nan")) == []
    assert wb.parse_pipe_values("") == []
    # numeric cell goes through as.character()
    assert wb.parse_pipe_values(1.0) == ["1"]


def test_parse_pairwise_values():
    assert wb.parse_pairwise_values("a~b|c~d") == [["a", "b"], ["c", "d"]]
    assert wb.parse_pairwise_values(None) == []
    with pytest.raises(ValueError, match="Invalid pairwise comparison entry: 'a'."):
        wb.parse_pairwise_values("a")


def test_auto_pairwise_values():
    assert wb.auto_pairwise_values(["a", "b", "c"]) == [["a", "b"], ["a", "c"], ["b", "c"]]
    assert wb.auto_pairwise_values(["only"]) == []
    assert wb.auto_pairwise_values([]) == []


def test_validate_unique_keys():
    ok = pd.DataFrame({"metric_key": ["a", "b"]})
    wb.validate_unique_keys(ok, "metric_key", "metrics")

    blank = pd.DataFrame({"metric_key": ["a", "  "]})
    with pytest.raises(ValueError, match="contains blank values in column 'metric_key'"):
        wb.validate_unique_keys(blank, "metric_key", "metrics")

    dup = pd.DataFrame({"metric_key": ["a", "b", "a"]})
    with pytest.raises(ValueError, match="has duplicate metric_key values: a"):
        wb.validate_unique_keys(dup, "metric_key", "metrics")


def test_normalize_workbook_tables_fills_all_sheets():
    normalized = wb.normalize_workbook_tables({})
    assert set(normalized) == set(wb.workbook_sheet_specs())
    assert list(normalized["metrics"].columns) == wb.workbook_sheet_columns()["metrics"]
    assert len(normalized["metrics"]) == 0
    # extras appended after the desired columns
    custom = wb.ensure_workbook_sheet_columns(
        pd.DataFrame({"extra": [1], "metric_key": ["m"]}), "metrics"
    )
    assert list(custom.columns)[:2] == ["metric_key", "display_name"]
    assert list(custom.columns)[-1] == "extra"


# --------------------------------------------------------------------------- #
# Function mappings (mirrors tests/function_mappings_sheet_tests.R)
# --------------------------------------------------------------------------- #

METRIC_KEYS = ["perRiffle", "WDR", "wBHR"]


def test_function_mappings_empty_returns_none():
    assert wb.build_function_mappings_from_tables({"function_mappings": pd.DataFrame()},
                                                  METRIC_KEYS) is None
    assert wb.build_function_mappings_from_tables({}, METRIC_KEYS) is None


def test_function_mappings_full_coverage():
    tables = _mapping_tables([
        {"discipline": "Geomorphology", "function_label": "Bedform Diversity",
         "metric_key": "perRiffle"},
        {"discipline": "Hydraulics", "function_label": "Width/Depth Ratio",
         "metric_key": "WDR"},
        {"discipline": "Hydraulics", "function_label": "Floodplain Connectivity",
         "metric_key": "wBHR"},
    ])
    out = wb.build_function_mappings_from_tables(tables, METRIC_KEYS)
    assert out is not None and len(out) == 3
    assert out.attrs["covers_all_metrics"] is True
    assert list(out.columns) == ["metric_key", "discipline", "function_label", "sort_order"]
    # Hydraulics precedes Geomorphology in the fixed discipline order;
    # within a discipline function_label sorts alphabetically.
    assert list(out["discipline"]) == ["Hydraulics", "Hydraulics", "Geomorphology"]
    assert list(out["function_label"]) == [
        "Floodplain Connectivity", "Width/Depth Ratio", "Bedform Diversity",
    ]
    assert list(out["sort_order"]) == [1, 2, 3]


def test_function_mappings_partial_coverage():
    tables = _mapping_tables([
        {"discipline": "Geomorphology", "function_label": "Bedform Diversity",
         "metric_key": "perRiffle"},
        {"discipline": "Hydraulics", "function_label": "Width/Depth Ratio",
         "metric_key": "WDR"},
    ])
    out = wb.build_function_mappings_from_tables(tables, METRIC_KEYS)
    assert out is not None and len(out) == 2
    assert out.attrs["covers_all_metrics"] is False


def test_function_mappings_unknown_metric_dropped(caplog):
    tables = _mapping_tables([
        {"discipline": "Geomorphology", "function_label": "Bedform Diversity",
         "metric_key": "perRiffle"},
        {"discipline": "Geomorphology", "function_label": "Phantom Fn",
         "metric_key": "someGhostMetric"},
    ])
    with caplog.at_level(logging.WARNING, logger="streamcurves"):
        out = wb.build_function_mappings_from_tables(tables, METRIC_KEYS)
    assert out is not None
    assert "someGhostMetric" not in list(out["metric_key"])
    assert out.attrs["covers_all_metrics"] is False
    assert any("someGhostMetric" in rec.message for rec in caplog.records)


def test_function_mappings_bad_discipline():
    tables = _mapping_tables([
        {"discipline": "Climatology", "function_label": "Snowpack",
         "metric_key": "perRiffle"},
    ])
    with pytest.raises(ValueError, match="unknown discipline"):
        wb.build_function_mappings_from_tables(tables, METRIC_KEYS)


def test_function_mappings_metric_reuse_allowed():
    tables = _mapping_tables([
        {"discipline": "Geomorphology", "function_label": "Bedform Diversity",
         "metric_key": "perRiffle"},
        {"discipline": "Geomorphology", "function_label": "Sediment continuity",
         "metric_key": "perRiffle"},
    ])
    out = wb.build_function_mappings_from_tables(tables, METRIC_KEYS)
    assert out is not None and len(out) == 2
    assert list(out["metric_key"]).count("perRiffle") == 2


def test_function_mappings_duplicate_same_function_case_insensitive():
    tables = _mapping_tables([
        {"discipline": "Geomorphology", "function_label": "Bedform Diversity",
         "metric_key": "perRiffle"},
        {"discipline": "Geomorphology", "function_label": "bedform diversity",
         "metric_key": "perRiffle"},
    ])
    with pytest.raises(ValueError, match="same function"):
        wb.build_function_mappings_from_tables(tables, METRIC_KEYS)


def test_function_mappings_cross_discipline_error():
    tables = _mapping_tables([
        {"discipline": "Geomorphology", "function_label": "Shared Fn",
         "metric_key": "perRiffle"},
        {"discipline": "Hydraulics", "function_label": "Shared Fn",
         "metric_key": "WDR"},
    ])
    with pytest.raises(ValueError, match="more than one discipline"):
        wb.build_function_mappings_from_tables(tables, METRIC_KEYS)


def test_function_mappings_bucket_rows():
    tables = _mapping_tables([
        {"discipline": "Geomorphology", "function_label": "Bedform Diversity",
         "metric_key": "perRiffle"},
        {"discipline": "Biology", "function_label": "Macroinvertebrate Community",
         "metric_key": ""},
        {"discipline": "Physicochemistry", "function_label": "Thermal Regime",
         "metric_key": None},
    ])
    out = wb.build_function_mappings_from_tables(tables, METRIC_KEYS)
    assert out is not None and len(out) == 3
    n_blank = sum(1 for v in out["metric_key"] if v is None or v == "")
    assert n_blank == 2
    assert "Biology" in list(out["discipline"])
    assert "Macroinvertebrate Community" in list(out["function_label"])


# --------------------------------------------------------------------------- #
# Synthetic workbook: happy path
# --------------------------------------------------------------------------- #


def test_read_input_workbook_happy_path(tmp_path):
    bundle = wb.read_input_workbook(_workbook_path(tmp_path))
    assert set(bundle) == BUNDLE_KEYS

    raw = bundle["raw_data"]
    assert len(raw) == 6
    assert list(raw[wb.SITE_ID_COL]) == [1, 2, 3, 4, 5, 6]
    # first data column is the default label source; blank cell -> "Site {id}"
    assert list(raw[wb.SITE_LABEL_COL]) == ["s1", "s2", "Site 3", "s4", "s5", "s6"]

    m_width = bundle["metric_config"]["m_width"]
    assert m_width["display_name"] == "Bankfull Width"
    assert m_width["column_name"] == "Width"
    assert m_width["units"] == "m"
    assert m_width["metric_family"] == "continuous"
    assert m_width["higher_is_better"] is True
    assert m_width["monotonic_linear"] is True
    assert m_width["preferred_transform"] == "log"
    assert m_width["min_sample_size"] == 10  # blank -> default
    assert m_width["best_subsets_allowed"] is True
    assert m_width["count_model"] is False
    assert m_width["stratification_mode"] == "subset"
    assert m_width["include_in_summary"] is True
    assert m_width["missing_data_rule"] is None
    assert m_width["notes"] == ""
    # metric_predictors sorted by sort_order (p_slope has 1, p_da has 2)
    assert m_width["allowed_predictors"] == ["p_slope", "p_da"]
    assert m_width["allowed_stratifications"] == ["s_eco"]

    m_wood = bundle["metric_config"]["m_wood"]
    assert m_wood["display_name"] == "m_wood"  # default = metric_key
    assert m_wood["units"] == ""
    assert m_wood["metric_family"] == "continuous"
    assert m_wood["higher_is_better"] is True  # numeric 1 -> "1" -> True
    assert m_wood["min_sample_size"] == 5
    assert m_wood["count_model"] is True
    assert m_wood["include_in_summary"] is False
    assert m_wood["missing_data_rule"] == "warn"
    assert m_wood["allowed_predictors"] == []

    s_eco = bundle["strat_config"]["s_eco"]
    assert s_eco["type"] == "single"
    assert s_eco["column_name"] == "Ecoregion"
    assert s_eco["levels"] == ["ECBP", "HELP", "IP"]  # sorted unique from data
    assert s_eco["pairwise_comparisons"] == [["ECBP", "HELP"], ["ECBP", "IP"], ["HELP", "IP"]]
    assert s_eco["min_group_size"] == 3

    s_size = bundle["strat_config"]["s_size"]
    assert s_size["is_custom_grouping"] is True
    assert s_size["column_name"] == "s_size"  # derived_column_name defaults to strat_key
    assert s_size["source_column"] == "Width"
    assert s_size["source_data_type"] == "continuous"
    assert s_size["min_group_size"] == 5  # blank -> default
    labels = [d["group_label"] for d in s_size["group_definitions"]]
    assert labels == ["Small", "Big"]  # arranged by numeric sort_order
    assert s_size["group_definitions"][0]["rule_expression"] == "<= 5"
    assert s_size["group_definitions"][0]["sort_order"] == 1.0
    assert s_size["levels"] == ["Small", "Big"]
    assert s_size["pairwise_comparisons"] == [["Small", "Big"]]

    s_pair = bundle["strat_config"]["s_pair"]
    assert s_pair["type"] == "paired"
    assert s_pair["column_name"] is None
    assert s_pair["primary"] == "s_eco"
    assert s_pair["secondary"] == "s_size"
    assert s_pair["levels"] == [] and s_pair["pairwise_comparisons"] == []

    s_recode = bundle["strat_config"]["s_recode"]
    assert s_recode["column_name"] == "EcoGroup"  # recode target column
    assert s_recode["levels"] == ["Agg", "Other"]  # collapse_map keys
    assert s_recode["pairwise_comparisons"] == [["Agg", "Other"]]

    p_da = bundle["predictor_config"]["p_da"]
    assert p_da["display_name"] == "Drainage area"
    assert p_da["column_name"] == "DA_mi2"
    assert p_da["type"] == "continuous"
    assert p_da["derived"] is False
    assert p_da["derivation_method"] == "none"
    assert p_da["source_columns"] == []
    assert math.isnan(p_da["constant"])
    assert p_da["expected_range"] == [0.0, 100.0]
    assert p_da["missing_data_rule"] == "error"

    p_slope = bundle["predictor_config"]["p_slope"]
    assert p_slope["display_name"] == "p_slope"  # default = predictor_key
    assert p_slope["derived"] is True
    assert p_slope["derivation_method"] == "ratio"
    assert p_slope["source_columns"] == ["A", "B"]
    assert p_slope["constant"] == 2.59
    assert all(math.isnan(v) for v in p_slope["expected_range"])
    assert p_slope["missing_data_rule"] == "warn"

    rec = bundle["factor_recode_config"]["rec_eco"]
    assert rec["source_column"] == "Ecoregion"
    assert rec["target_column"] == "EcoGroup"
    assert rec["collapse_map"] == {"Agg": ["ECBP", "HELP"], "Other": ["IP"]}

    mask = bundle["site_mask_config"]
    assert mask == {"site_label_column": "SiteName", "masked_site_ids": [], "site_labels": []}

    assert bundle["discipline_function_mapping"] is None
    assert bundle["mapping_covers_all_metrics"] is False

    metadata = bundle["metadata"]
    assert set(metadata) == set(wb.workbook_sheet_specs())
    assert len(metadata["data"]) == 6
    assert len(metadata["site_masks"]) == 0
    assert list(metadata["site_mask_settings"]["site_label_column"]) == ["SiteName"]


# --------------------------------------------------------------------------- #
# Synthetic workbook: validation errors
# --------------------------------------------------------------------------- #


def test_missing_required_sheet(tmp_path):
    sheets = _base_sheets()
    del sheets["predictors"]
    with pytest.raises(ValueError, match="Workbook is missing required sheets: predictors"):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_missing_required_columns(tmp_path):
    sheets = _base_sheets()
    header = sheets["metrics"][0]
    drop = {"metric_family", "higher_is_better"}
    keep_idx = [i for i, c in enumerate(header) if c not in drop]
    sheets["metrics"] = [[row[i] for i in keep_idx] for row in sheets["metrics"]]
    with pytest.raises(
        ValueError,
        match="Sheet 'metrics' is missing required columns: metric_family, higher_is_better",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_bad_flag_value(tmp_path):
    sheets = _base_sheets()
    sheets["metrics"][1][5] = "maybe"  # higher_is_better
    with pytest.raises(ValueError, match="Could not parse logical value 'maybe'."):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_duplicate_metric_keys(tmp_path):
    sheets = _base_sheets()
    sheets["metrics"].append(list(sheets["metrics"][1]))
    with pytest.raises(ValueError, match="Sheet 'metrics' has duplicate metric_key values: m_width"):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_blank_strat_key(tmp_path):
    sheets = _base_sheets()
    sheets["stratifications"][1][0] = None
    with pytest.raises(
        ValueError, match="Sheet 'stratifications' contains blank values in column 'strat_key'."
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_unknown_foreign_key(tmp_path):
    sheets = _base_sheets()
    sheets["metric_predictors"].append(["m_width", "p_ghost", 3])
    with pytest.raises(
        ValueError,
        match="Sheet 'metric_predictors' contains unknown metric_key or predictor_key values.",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_bad_strat_type(tmp_path):
    sheets = _base_sheets()
    sheets["stratifications"][1][2] = "diagonal"
    with pytest.raises(
        ValueError, match="Unsupported strat_type 'diagonal' for stratification 's_eco'."
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_blank_strat_type_quirk(tmp_path):
    # NOTE(parity): R's compact_chr(...)[1] %||% stop() never fires the stop
    # (indexing yields NA, not NULL), so a blank strat_type reports
    # "Unsupported strat_type 'NA'" rather than "is missing strat_type".
    sheets = _base_sheets()
    sheets["stratifications"][1][2] = None
    with pytest.raises(
        ValueError, match="Unsupported strat_type 'NA' for stratification 's_eco'."
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_paired_missing_keys(tmp_path):
    sheets = _base_sheets()
    sheets["stratifications"][3][6] = None  # drop secondary_strat_key
    with pytest.raises(
        ValueError,
        match="Paired stratification 's_pair' must provide primary_strat_key and secondary_strat_key.",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_paired_unknown_base_keys(tmp_path):
    sheets = _base_sheets()
    sheets["stratifications"][3][5] = "s_ghost"  # primary_strat_key
    with pytest.raises(
        ValueError,
        match="Paired stratifications reference unknown base stratification keys: s_pair",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_raw_single_missing_source_column(tmp_path):
    sheets = _base_sheets()
    sheets["stratifications"][1][3] = "NotAColumn"
    with pytest.raises(
        ValueError,
        match="Raw stratification 's_eco' references missing data column 'NotAColumn'.",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_custom_group_requires_rows(tmp_path):
    sheets = _base_sheets()
    sheets["strat_groups"] = [sheets["strat_groups"][0]]  # header only
    with pytest.raises(
        ValueError, match="Custom grouping 's_size' has no rows in sheet 'strat_groups'."
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_custom_group_bad_source_data_type(tmp_path):
    sheets = _base_sheets()
    sheets["stratifications"][2][4] = "fuzzy"
    with pytest.raises(
        ValueError,
        match="Custom grouping 's_size' must set source_data_type to 'categorical' or 'continuous'.",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_custom_group_derived_column_collision(tmp_path):
    sheets = _base_sheets()
    sheets["stratifications"][2][7] = "Wood"  # existing data column
    with pytest.raises(
        ValueError,
        match="Custom grouping 's_size' would overwrite existing data column 'Wood'.",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_custom_group_duplicate_group_labels(tmp_path):
    sheets = _base_sheets()
    sheets["strat_groups"].append(["s_size", "Big", 3, None, "> 7"])
    with pytest.raises(
        ValueError, match="Custom grouping 's_size' has duplicate group_label values."
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_read_input_workbook_not_found(tmp_path):
    with pytest.raises(ValueError, match="Input workbook not found"):
        wb.read_input_workbook(tmp_path / "nope.xlsx")


def test_read_input_workbook_requires_xlsx(tmp_path):
    path = tmp_path / "input.csv"
    path.write_text("x", encoding="utf-8")
    with pytest.raises(ValueError, match=r"Input workbook must be an \.xlsx file\."):
        wb.read_input_workbook(path)


# --------------------------------------------------------------------------- #
# Site masks
# --------------------------------------------------------------------------- #


def _mask_sheets(masked=(2, 5), label_col="SiteName"):
    sheets = _base_sheets()
    sheets["site_masks"] = [["masked_sites", "site_label"]] + [[m, None] for m in masked]
    sheets["site_mask_settings"] = [["site_label_column"], [label_col]]
    return sheets


def test_site_masks_applied(tmp_path):
    bundle = wb.read_input_workbook(_workbook_path(tmp_path, _mask_sheets()))
    raw = bundle["raw_data"]
    assert len(raw) == 4
    assert list(raw[wb.SITE_ID_COL]) == [1, 3, 4, 6]

    mask = bundle["site_mask_config"]
    assert mask["site_label_column"] == "SiteName"
    assert mask["masked_site_ids"] == [2, 5]
    assert mask["site_labels"] == ["s2", "s5"]

    # metadata carries the canonical rebuilt mask tables
    masks_tbl = bundle["metadata"]["site_masks"]
    assert list(masks_tbl["masked_sites"]) == [2, 5]
    assert list(masks_tbl["site_label"]) == ["s2", "s5"]
    assert list(bundle["metadata"]["site_mask_settings"]["site_label_column"]) == ["SiteName"]


def test_site_masks_invalid_ids(tmp_path):
    with pytest.raises(
        ValueError, match="Sheet 'site_masks' contains invalid masked_sites values: 99"
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, _mask_sheets(masked=(2, 99))))


def test_site_masks_duplicate_ids(tmp_path):
    with pytest.raises(
        ValueError, match="Sheet 'site_masks' contains duplicate masked_sites values: 2"
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, _mask_sheets(masked=(2, 2))))


def test_site_mask_settings_multiple_values(tmp_path):
    sheets = _mask_sheets()
    sheets["site_mask_settings"].append(["Ecoregion"])
    with pytest.raises(
        ValueError,
        match="Sheet 'site_mask_settings' must contain at most one distinct site_label_column value.",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, sheets))


def test_site_mask_settings_missing_column(tmp_path):
    with pytest.raises(
        ValueError,
        match="Sheet 'site_mask_settings' references missing site_label_column 'Nope'.",
    ):
        wb.read_input_workbook(_workbook_path(tmp_path, _mask_sheets(label_col="Nope")))


# --------------------------------------------------------------------------- #
# OSAM fixture: smoke + round trip + golden
# --------------------------------------------------------------------------- #


def _require_fixture():
    if not FIXTURE.exists():
        pytest.skip(f"fixture missing: {FIXTURE}")


def test_smoke_osam_fixture():
    _require_fixture()
    bundle = wb.load_data(FIXTURE)
    assert set(bundle) == BUNDLE_KEYS

    raw = bundle["raw_data"]
    assert len(raw) == 39
    assert wb.SITE_ID_COL in raw.columns
    assert wb.SITE_LABEL_COL in raw.columns
    assert list(raw[wb.SITE_ID_COL]) == list(range(1, 40))

    assert len(bundle["metric_config"]) == 22
    assert "perRiffle" in bundle["metric_config"]
    assert bundle["metric_config"]["perRiffle"]["column_name"]

    assert len(bundle["strat_config"]) == 11
    assert bundle["strat_config"]["Ecoregion_x_DACAT"]["type"] == "paired"
    assert bundle["strat_config"]["Ecoregion_grouped"]["is_custom_grouping"] is True

    assert len(bundle["predictor_config"]) == 16
    assert set(bundle["factor_recode_config"]) == {"StreamType2", "BEHIcombined", "BEHIeroding"}

    assert bundle["site_mask_config"]["site_label_column"] == "ID"
    assert bundle["site_mask_config"]["masked_site_ids"] == []

    mapping = bundle["discipline_function_mapping"]
    assert mapping is not None and len(mapping) == 22
    assert bundle["mapping_covers_all_metrics"] is True


def _equalish(a, b) -> bool:
    if isinstance(a, dict) and isinstance(b, dict):
        return a.keys() == b.keys() and all(_equalish(a[k], b[k]) for k in a)
    if isinstance(a, (list, tuple)) and isinstance(b, (list, tuple)):
        return len(a) == len(b) and all(_equalish(x, y) for x, y in zip(a, b))
    if isinstance(a, float) and isinstance(b, float):
        return (math.isnan(a) and math.isnan(b)) or a == b
    return a == b


def _assert_bundles_equal(b1: dict, b2: dict) -> None:
    assert set(b1) == set(b2)
    pd.testing.assert_frame_equal(b1["raw_data"], b2["raw_data"], check_dtype=False)
    for key in ("metric_config", "strat_config", "predictor_config",
                "factor_recode_config", "site_mask_config"):
        assert _equalish(b1[key], b2[key]), f"bundle piece differs: {key}"
    m1, m2 = b1["discipline_function_mapping"], b2["discipline_function_mapping"]
    assert (m1 is None) == (m2 is None)
    if m1 is not None:
        pd.testing.assert_frame_equal(m1, m2, check_dtype=False)
    assert b1["mapping_covers_all_metrics"] == b2["mapping_covers_all_metrics"]
    assert set(b1["metadata"]) == set(b2["metadata"])
    for sheet_name in b1["metadata"]:
        pd.testing.assert_frame_equal(
            b1["metadata"][sheet_name],
            b2["metadata"][sheet_name],
            check_dtype=False,
        )


def test_round_trip_synthetic(tmp_path):
    bundle1 = wb.read_input_workbook(_workbook_path(tmp_path))
    out = tmp_path / "rewritten.xlsx"
    wb.write_input_workbook(bundle1["metadata"], out)
    bundle2 = wb.read_input_workbook(out)
    _assert_bundles_equal(bundle1, bundle2)


def test_round_trip_osam(tmp_path):
    _require_fixture()
    bundle1 = wb.read_input_workbook(FIXTURE)
    out = tmp_path / "osam_rewritten.xlsx"
    wb.write_input_workbook(bundle1["metadata"], out)
    bundle2 = wb.read_input_workbook(out)
    _assert_bundles_equal(bundle1, bundle2)


def test_write_to_binary_stream(tmp_path):
    """The Save Workbook download handler passes an io.BytesIO buffer, not a
    path (views/data_overview.py download_workbook)."""
    import io

    bundle1 = wb.read_input_workbook(_workbook_path(tmp_path))
    buf = io.BytesIO()
    wb.write_input_workbook(bundle1["metadata"], buf)
    data = buf.getvalue()
    assert data[:2] == b"PK"  # a real xlsx (zip) was written to the stream
    # and it round-trips back through the reader
    reread = tmp_path / "from_stream.xlsx"
    reread.write_bytes(data)
    _assert_bundles_equal(bundle1, wb.read_input_workbook(reread))


def _norm_py(v):
    """Normalize Python bundle values for golden comparison (nan -> None,
    numpy scalars -> Python scalars)."""
    if isinstance(v, np.generic):
        v = v.item()
    if isinstance(v, float) and math.isnan(v):
        return None
    if isinstance(v, dict):
        return {k: _norm_py(x) for k, x in v.items()}
    if isinstance(v, (list, tuple)):
        return [_norm_py(x) for x in v]
    return v


def _assert_matches_golden(py, golden, path: str) -> None:
    """Compare against R jsonlite output: R scalars arrive boxed as length-1
    arrays ([true], ["ID"], [null]); R NULL arrives as unboxed null."""
    py = _norm_py(py)
    if isinstance(golden, dict):
        assert isinstance(py, dict), f"{path}: expected dict, got {type(py).__name__}"
        assert set(py.keys()) == set(golden.keys()), (
            f"{path}: key mismatch (extra: {set(py) - set(golden)}, "
            f"missing: {set(golden) - set(py)})"
        )
        for k in golden:
            _assert_matches_golden(py[k], golden[k], f"{path}.{k}")
        return
    if isinstance(golden, list):
        if isinstance(py, list):
            assert len(py) == len(golden), (
                f"{path}: length {len(py)} != golden {len(golden)}"
            )
            for i, (a, b) in enumerate(zip(py, golden)):
                _assert_matches_golden(a, b, f"{path}[{i}]")
            return
        # jsonlite boxes R scalars as length-1 arrays.
        assert len(golden) == 1, f"{path}: scalar {py!r} vs golden array {golden!r}"
        _assert_matches_golden(py, golden[0], path)
        return
    assert py == golden, f"{path}: {py!r} != golden {golden!r}"


def test_golden_bundle_meta():
    if not GOLDEN.exists():
        pytest.skip(f"golden fixture missing: {GOLDEN}")
    _require_fixture()

    golden = json.loads(GOLDEN.read_text(encoding="utf-8"))
    bundle = wb.read_input_workbook(FIXTURE)
    raw = bundle["raw_data"]
    mapping = bundle["discipline_function_mapping"]
    actual = {
        "metric_config": bundle["metric_config"],
        "strat_config": bundle["strat_config"],
        "predictor_config": bundle["predictor_config"],
        "factor_recode_config": bundle["factor_recode_config"],
        "site_mask_config": bundle["site_mask_config"],
        # jsonlite serializes the mapping tibble column-wise
        "discipline_function_mapping": None
        if mapping is None
        else {col: list(mapping[col]) for col in mapping.columns},
        "covers_all_metrics": bundle["mapping_covers_all_metrics"],
        "raw_data_dim": [len(raw), int(raw.shape[1])],
        "raw_data_names": [str(c) for c in raw.columns],
    }
    checked = 0
    for key, expected in golden.items():
        if key in actual:
            _assert_matches_golden(actual[key], expected, key)
            checked += 1
    assert checked >= 5, "golden file had too few comparable keys"
